"""Update the tracking Excel workbook with projected and observed ROI."""

from __future__ import annotations

import argparse
import json
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hippique_orchestrator.post_course_payload import (
    JsonDict,
    PostCourseSummary,
    build_payload_from_sources,
    merge_meta,
    summarise_ticket_metrics,
)

PREVISION_HEADERS = [
    "R/C",
    "hippodrome",
    "date",
    "discipline",
    "mises",
    "EV_sp",
    "EV_global",
    "ROI_sp",
    "ROI_global",
    "risk_of_ruin",
    "clv_moyen",
    "variance",
    "payout_attendu",
    "model",
]


OBSERVED_HEADERS = [
    "R/C",
    "hippodrome",
    "date",
    "discipline",
    "mises",
    "gains",
    "ROI_reel",
    "result_moyen",
    "ROI_reel_moyen",
    "Brier_total",
    "Brier_moyen",
    "EV_total",
    "EV_ecart",
    "model",
]


SUIVI_HEADERS = [
    "R/C",
    "date",
    "hippodrome",
    "discipline",
    "mises",
    "gains",
    "ROI_reel",
    "ROI_estime",
    "payout_attendu",
    "verdict",
    "notes",
]


def _load_json(path: str | Path) -> JsonDict:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def _load_workbook(path: Path) -> tuple[Workbook, bool]:
    if path.exists():
        return load_workbook(path), False
    return Workbook(), True


def _get_sheet(wb: Workbook, title: str) -> Worksheet:
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title=title)


def _ensure_header_map(ws: Worksheet, headers: Iterable[str]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    max_column = ws.max_column
    blank_sheet = (
        ws.max_row <= 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value in (None, "")
    )
    if not blank_sheet:
        for col in range(1, max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value not in (None, ""):
                header_map[str(value)] = col
    next_col = 0 if blank_sheet else max_column
    for header in headers:
        if header not in header_map:
            next_col += 1
            ws.cell(row=1, column=next_col, value=header)
            header_map[header] = next_col
    return header_map


def _upsert_row(
    ws: Worksheet,
    headers: Iterable[str],
    values: dict[str, Any],
    *,
    key_header: str = "R/C",
) -> int | None:
    header_map = _ensure_header_map(ws, headers)
    key_value = values.get(key_header)
    if key_value in (None, ""):
        return
    key_col = header_map[key_header]
    key_str = str(key_value)
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=key_col).value
        if cell_value is None:
            continue
        if str(cell_value) == key_str:
            target_row = row_idx
            break
    if target_row is None:
        target_row = max(ws.max_row + 1, 2)
    for header in headers:
        col_idx = header_map[header]
        ws.cell(row=target_row, column=col_idx, value=values.get(header))
    return target_row


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _update_previsionnel_sheet(
    wb: Workbook,
    sheet_name: str,
    meta: JsonDict,
    *,
    total_stake: float,
    ev_data: JsonDict | None,
) -> None:
    rc = meta.get("rc")
    if not rc:
        return
    ws = _get_sheet(wb, sheet_name)
    ev_section = ev_data or {}
    row = {
        "R/C": rc,
        "hippodrome": meta.get("hippodrome", ""),
        "date": meta.get("date", ""),
        "discipline": meta.get("discipline", ""),
        "mises": round(total_stake, 2) if total_stake else 0.0,
        "EV_sp": ev_section.get("sp"),
        "EV_global": ev_section.get("global"),
        "ROI_sp": ev_section.get("roi_sp"),
        "ROI_global": ev_section.get("roi_global"),
        "risk_of_ruin": ev_section.get("risk_of_ruin"),
        "clv_moyen": ev_section.get("clv_moyen"),
        "variance": ev_section.get("variance"),
        "payout_attendu": ev_section.get("combined_expected_payout"),
        "model": meta.get("model") or meta.get("MODEL", ""),
    }
    _upsert_row(ws, PREVISION_HEADERS, row)


def _update_observe_sheet(
    wb: Workbook,
    sheet_name: str,
    meta: JsonDict,
    *,
    tickets: list[JsonDict],
    summary: PostCourseSummary,
    observed: JsonDict | None,
    total_stake: float,
    total_gain: float,
) -> None:
    rc = meta.get("rc")
    if not rc:
        return
    observed_section = observed or {}
    roi_observed = observed_section.get("roi_reel", summary.roi)
    if roi_observed is None and total_stake:
        roi_observed = (total_gain - total_stake) / total_stake if total_stake else 0.0
    has_observed_metrics = any(
        (
            roi_observed,
            total_gain,
            any("gain_reel" in t for t in tickets),
        )
    )
    if not has_observed_metrics:
        return

    row = {
        "R/C": rc,
        "hippodrome": meta.get("hippodrome", ""),
        "date": meta.get("date", ""),
        "discipline": meta.get("discipline", ""),
        "mises": round(total_stake, 2) if total_stake else 0.0,
        "gains": round(total_gain, 2) if total_gain else 0.0,
        "ROI_reel": roi_observed,
        "result_moyen": observed_section.get("result_moyen", summary.result_mean),
        "ROI_reel_moyen": observed_section.get("roi_reel_moyen", summary.roi_ticket_mean),
        "Brier_total": observed_section.get("brier_total", summary.brier_total),
        "Brier_moyen": observed_section.get("brier_moyen", summary.brier_mean),
        "EV_total": observed_section.get("ev_total", summary.ev_total),
        "EV_ecart": observed_section.get("ev_ecart_total", summary.ev_diff_total),
        "model": meta.get("model") or meta.get("MODEL", ""),
    }
    ws = _get_sheet(wb, sheet_name)
    _upsert_row(ws, OBSERVED_HEADERS, row)


def _normalise_notes(notes: Any) -> str:
    if isinstance(notes, (list, tuple, set)):
        return "; ".join(str(item) for item in notes if str(item).strip())
    if notes in (None, ""):
        return ""
    return str(notes)


def _update_suivi_sheet(
    wb: Workbook,
    sheet_name: str,
    meta: JsonDict,
    *,
    total_stake: float,
    total_gain: float,
    ev_data: JsonDict | None,
    observed: JsonDict | None,
    payload: JsonDict,
) -> tuple[Worksheet, int | None]:
    rc = meta.get("rc")
    if not rc:
        return _get_sheet(wb, sheet_name), None

    ws = _get_sheet(wb, sheet_name)
    ev_section = ev_data or {}
    observed_section = observed or {}
    roi_reel = 0.0
    if total_stake:
        roi_reel = (total_gain - total_stake) / total_stake
    roi_estime = ev_section.get("roi_global")
    payout = ev_section.get("combined_expected_payout")
    row = {
        "R/C": rc,
        "date": meta.get("date", ""),
        "hippodrome": meta.get("hippodrome", ""),
        "discipline": meta.get("discipline", ""),
        "mises": round(total_stake, 2) if total_stake else 0.0,
        "gains": round(total_gain, 2) if total_gain else 0.0,
        "ROI_reel": roi_reel,
        "ROI_estime": roi_estime,
        "payout_attendu": payout,
        "verdict": payload.get("verdict") or observed_section.get("verdict"),
        "notes": _normalise_notes(payload.get("notes") or observed_section.get("notes")),
    }
    row_idx = _upsert_row(ws, SUIVI_HEADERS, row)
    return ws, row_idx


def _collect_row(ws: Worksheet, headers: Iterable[str], row_idx: int) -> dict[str, Any]:
    header_map = _ensure_header_map(ws, headers)
    result: dict[str, Any] = {}
    for header in headers:
        col_idx = header_map[header]
        result[header] = ws.cell(row=row_idx, column=col_idx).value
    return result


def _print_row(prefix: str, row: Mapping[str, Any]) -> None:
    printable: dict[str, Any] = {}
    for key, value in row.items():
        if value in (None, ""):
            continue
        if isinstance(value, float):
            printable[key] = round(value, 6)
        else:
            printable[key] = value
    if not printable:
        return
    print(f"{prefix} {json.dumps(printable, ensure_ascii=False, sort_keys=True)}")


def update_excel(excel_path_str: str, payload_path_str: str | None = None, arrivee_path_str: str | None = None, tickets_path_str: str | None = None, sheet_prevision: str = "ROI Prévisionnel", sheet_observe: str = "ROI Observé") -> None:
    """Updates the tracking Excel workbook with race results."""
    excel_path = Path(excel_path_str)

    if not payload_path_str and (not arrivee_path_str or not tickets_path_str):
        raise ValueError("--payload or the duo --arrivee/--tickets must be provided")

    if payload_path_str:
        payload_path = Path(payload_path_str)
        payload = _load_json(payload_path)
    else:
        arrivee_path = Path(arrivee_path_str)
        tickets_path = Path(tickets_path_str)
        arrivee_data = _load_json(arrivee_path)
        tickets_data = _load_json(tickets_path)
        payload = build_payload_from_sources(arrivee_data, tickets_data)

    if not isinstance(payload, dict):
        raise ValueError("Le payload post-course doit être un objet JSON.")

    meta_source = payload.get("meta", {}) if isinstance(payload, dict) else {}
    payload_arrivee = payload.get("arrivee") if isinstance(payload, dict) else {}
    meta = merge_meta(payload_arrivee or {}, {"meta": meta_source})

    raw_tickets = payload.get("tickets")
    tickets_list = (
        [t for t in raw_tickets if isinstance(t, dict)] if isinstance(raw_tickets, list) else []
    )
    summary = summarise_ticket_metrics(tickets_list)

    mises_raw = payload.get("mises")
    mises_section = mises_raw if isinstance(mises_raw, dict) else {}
    total_stake_value = mises_section.get("total") if mises_section else None
    if total_stake_value is None and mises_section:
        total_stake_value = mises_section.get("totales")
    total_stake = _as_float(total_stake_value, summary.total_stake)
    total_gain = _as_float(
        mises_section.get("gains") if mises_section else None, summary.total_gain
    )

    ev_raw = payload.get("ev_estimees")
    ev_data = ev_raw if isinstance(ev_raw, dict) else {}
    observed_raw = payload.get("ev_observees")
    observed = observed_raw if isinstance(observed_raw, dict) else {}

    wb, created = _load_workbook(excel_path)
    if created and "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        default_ws = wb["Sheet"]
        if default_ws.max_row <= 1 and default_ws.cell(row=1, column=1).value in (
            None,
            "",
        ):
            wb.remove(default_ws)

    _update_previsionnel_sheet(
        wb,
        sheet_prevision,
        meta,
        total_stake=total_stake,
        ev_data=ev_data,
    )
    _update_observe_sheet(
        wb,
        sheet_observe,
        meta,
        tickets=tickets_list,
        summary=summary,
        observed=observed,
        total_stake=total_stake,
        total_gain=total_gain,
    )

    ws_suivi, row_idx = _update_suivi_sheet(
        wb,
        "Suivi",
        meta,
        total_stake=total_stake,
        total_gain=total_gain,
        ev_data=ev_data,
        observed=observed,
        payload=payload,
    )
    if row_idx:
        row_data = _collect_row(ws_suivi, SUIVI_HEADERS, row_idx)
        _print_row("Suivi:", row_data)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)

def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Mettre à jour le ROI dans l'Excel de suivi")
    parser.add_argument(
        "--excel",
        default="modele_suivi_courses_hippiques.xlsx",
        help="Chemin du classeur Excel à mettre à jour",
    )
    parser.add_argument("--payload", help="JSON normalisé généré après le post-course")
    parser.add_argument("--arrivee", help="JSON de l'arrivée officielle")
    parser.add_argument("--tickets", help="JSON des tickets (p_finale.json)")
    parser.add_argument(
        "--sheet-prevision",
        default="ROI Prévisionnel",
        help="Nom de la feuille pour les projections",
    )
    parser.add_argument(
        "--sheet-observe",
        default="ROI Observé",
        help="Nom de la feuille pour le ROI observé",
    )
    args = parser.parse_args(argv)

    try:
        update_excel(
            excel_path_str=args.excel,
            payload_path_str=args.payload,
            arrivee_path_str=args.arrivee,
            tickets_path_str=args.tickets,
            sheet_prevision=args.sheet_prevision,
            sheet_observe=args.sheet_observe,
        )
    except (ValueError, FileNotFoundError) as e:
        parser.error(str(e))


if __name__ == "__main__":  # pragma: no cover
    main()
