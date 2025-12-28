"Utilities for synchronising the planning worksheet with pipeline outputs."

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
from collections.abc import Iterable, Mapping, Sequence
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hippique_orchestrator import config

try:  # pragma: no cover - fallback for older Python versions
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover - very defensive
    ZoneInfo = None  # type: ignore[assignment]


PLANNING_HEADERS: Sequence[str] = (
    "Date",
    "Réunion",
    "Course",
    "Hippodrome",
    "Heure",
    "Partants",
    "Discipline",
    "Statut H-30",
    "Statut H-5",
    "Jouable H-5",
    "Tickets H-5",
    "Commentaires",
)


@lru_cache(maxsize=1)
def _env_timezone() -> dt.tzinfo | None:
    """Return the timezone configured via ``$TZ`` when available."""

    if ZoneInfo is None:
        return None

    tz_name = config.TIMEZONE
    if not tz_name:
        return None

    try:
        return ZoneInfo(tz_name)
    except Exception:  # pragma: no cover - invalid/unknown TZ identifiers
        return None


def _phase_argument(value: str) -> str:
    cleaned = value.strip().upper().replace("-", "")
    if cleaned not in {"H30", "H5"}:
        raise argparse.ArgumentTypeError("phase must be H30 or H5")
    return cleaned


def _load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def _load_workbook(path: Path) -> tuple[Workbook, bool]:
    if path.exists():
        return load_workbook(path), False
    return Workbook(), True


def _prepare_sheet(wb: Workbook, title: str) -> Worksheet:
    if title in wb.sheetnames:
        return wb[title]
    if len(wb.sheetnames) == 1:
        ws = wb.active
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value in (None, ""):
            ws.title = title
            return ws
    return wb.create_sheet(title=title)


def _ensure_headers(ws: Worksheet, headers: Sequence[str]) -> Mapping[str, int]:
    header_map: dict[str, int] = {}
    blank_sheet = (
        ws.max_row <= 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value in (None, "")
    )
    max_col = 0 if blank_sheet else ws.max_column
    if not blank_sheet:
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value not in (None, ""):
                header_map[str(value)] = col
    next_col = max_col
    for header in headers:
        if header not in header_map:
            next_col += 1
            ws.cell(row=1, column=next_col, value=header)
            header_map[header] = next_col
    return header_map


def _coerce_number(value: Any) -> str:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    if abs(num - round(num)) < 1e-9:
        return str(int(round(num)))
    return f"{num:.2f}".rstrip("0").rstrip(".")


def _format_time(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\u202f", " ")
    if len(text) == 4 and text.isdigit():
        return f"{text[:2]}:{text[2:]}"
    if len(text) >= 5 and text[2] == ":" and text[:2].isdigit():
        try:
            hour = int(text[:2]) % 24
            minute = int(text[3:5])
        except ValueError:
            pass
        else:
            return f"{hour:02d}:{minute:02d}"
    cleaned = text.replace("Z", "+00:00")
    try:
        parsed = dt.datetime.fromisoformat(cleaned)
    except ValueError:
        time_pattern = re.compile(
            r"(\d{1,2})\s*(?:heures?|heure|hours?|hrs?|hres?|[hH:.])\s*(\d{1,2})?",
            re.IGNORECASE,
        )
        match = time_pattern.search(text)
        if match:
            hour = int(match.group(1)) % 24
            minute_str = match.group(2)
            minute = int(minute_str) if minute_str is not None else 0
            return f"{hour:02d}:{minute:02d}"
        hour_only = re.search(
            r"(\d{1,2})\s*(?:heures?|heure|hours?|hrs?|hres?|[hH])", text, re.IGNORECASE
        )
        if hour_only:
            hour = int(hour_only.group(1)) % 24
            return f"{hour:02d}:00"
        return text
    if parsed.tzinfo is not None:
        target_tz = _env_timezone()
        if target_tz is not None:
            parsed = parsed.astimezone(target_tz)
    return parsed.strftime("%H:%M")


def _first_non_empty(*values: Any) -> Any:
    for value in values:
        if isinstance(value, str):
            trimmed = value.strip()
            if trimmed:
                return trimmed
        elif value not in (None, ""):
            return value
    return None


def _blank_if_missing(value: Any) -> Any:
    """Return ``None`` for missing values to preserve existing cells."""

    if value in (None, ""):
        return None
    return value


def _format_row_identifier(row: Mapping[str, Any]) -> str:
    """Return a compact label describing ``row`` for CLI feedback."""

    def _clean(value: Any) -> str:
        if value in (None, ""):
            return ""
        return str(value).strip()

    reunion = _clean(row.get("Réunion"))
    course = _clean(row.get("Course"))
    rc = ""
    if reunion and course:
        rc = f"{reunion}{course}"
    else:
        rc = reunion or course
    if not rc:
        rc = _clean(row.get("RC")) or _clean(row.get("rc"))

    date = _clean(row.get("Date"))

    if rc and date:
        return f"{rc} – {date}"
    return rc or date


def _extract_common_meta(
    payload: Mapping[str, Any], parents: Sequence[Mapping[str, Any]] | None = None
) -> dict[str, Any]:
    sources: list[Mapping[str, Any]] = [payload]
    if parents:
        sources.extend(parent for parent in parents if isinstance(parent, Mapping))

    def _values(meta_keys: Sequence[str], payload_keys: Sequence[str]) -> list[Any]:
        candidates: list[Any] = []
        for source in sources:
            meta = source.get("meta") if isinstance(source.get("meta"), Mapping) else None
            if isinstance(meta, Mapping):
                candidates.extend(meta.get(key) for key in meta_keys)
            candidates.extend(source.get(key) for key in payload_keys)
        return candidates

    date = _first_non_empty(*_values(["date"], ["date"]))
    hippodrome = _first_non_empty(*_values(["hippodrome", "hippo"], ["hippodrome", "hippo"]))
    rc_value = _first_non_empty(*_values(["rc"], ["rc", "rc_label"]))
    reunion = _first_non_empty(
        *_values(["reunion", "meeting", "r"], ["reunion", "meeting", "r", "r_label"])
    )
    course = _first_non_empty(
        *_values(["course", "race", "c"], ["course", "race", "c", "course_label"])
    )
    if not reunion and isinstance(rc_value, str):
        match = re.search(r"R\s*(\d+)", rc_value, re.IGNORECASE)
        if match:
            reunion = f"R{int(match.group(1))}"
    if not course and isinstance(rc_value, str):
        match = re.search(r"C\s*(\d+)", rc_value, re.IGNORECASE)
        if match:
            course = f"C{int(match.group(1))}"
    partant_keys = [
        "partants",
        "nb_partants",
        "runners_count",
        "nb_partants_declares",
        "nombre_partants",
        "nb_participants",
        "nombre_participants",
        "participants_count",
    ]
    partants_val = _first_non_empty(*_values(partant_keys, partant_keys))
    partants = None
    if isinstance(partants_val, str):
        match = re.search(r"\d+", partants_val)
        if match:
            partants = int(match.group(0))
    elif partants_val is not None:
        try:
            partants = int(partants_val)
        except (ValueError, TypeError):
            pass

    if partants in (None, ""):
        for source in sources:
            meta = source.get("meta") if isinstance(source.get("meta"), Mapping) else None
            runners = None
            if isinstance(meta, Mapping):
                runners = meta.get("runners")
            if not runners:
                runners = source.get("runners")
            if isinstance(runners, Iterable) and not isinstance(runners, (str, bytes, Mapping)):
                partants = sum(1 for _ in runners)
                if partants not in (None, ""):
                    break
    start_time_keys = [
        "start_time",
        "start",
        "heure",
        "time",
        "hour",
        "startTime",
        "heure_depart",
        "heure_depart_programme",
        "heure_programme",
        "official_start_time",
        "official_time",
        "heure_officielle",
        "horaire",
    ]
    start_time = _first_non_empty(*_values(start_time_keys, start_time_keys))
    formatted_time = _format_time(start_time)
    discipline_keys = ["discipline", "type", "specialite", "speciality"]
    discipline = _first_non_empty(*_values(discipline_keys, discipline_keys))
    return {
        "date": date,
        "hippodrome": hippodrome,
        "reunion": reunion,
        "course": course,
        "start_time": formatted_time,
        "partants": partants,
        "discipline": discipline,
    }


def _course_like(obj: Mapping[str, Any]) -> bool:
    keys = set(obj.keys())
    meta = obj.get("meta") if isinstance(obj.get("meta"), Mapping) else None
    if isinstance(meta, Mapping):
        keys.update(meta.keys())
    hints = {
        "course",
        "race",
        "c",
        "course_label",
        "r_label",
        "num_course",
        "id_course",
        "numero_course",
        "event_id",
    }
    details = {
        "runners",
        "participants",
        "partants",
        "nb_participants",
        "nombre_participants",
        "participants_count",
        "horses",
        "odds",
        "start_time",
        "official_start_time",
        "official_time",
        "start",
        "heure",
        "time",
        "hour",
        "discipline",
    }
    has_hint = any(key in keys for key in hints)
    if not has_hint:
        return False
    if any(key in obj for key in details):
        return True
    if isinstance(meta, Mapping) and any(key in meta for key in details):
        return True
    return False


def _extract_course_payloads(
    payload: Mapping[str, Any],
) -> list[tuple[Mapping[str, Any], tuple[Mapping[str, Any], ...]]]:
    results: list[tuple[Mapping[str, Any], tuple[Mapping[str, Any], ...]]] = []
    visited: set[int] = set()

    def _walk(value: Any, parents: tuple[Mapping[str, Any], ...]) -> None:
        if isinstance(value, Mapping):
            obj_id = id(value)
            if obj_id in visited:
                return
            visited.add(obj_id)
            if value is not payload and _course_like(value):
                results.append((value, parents))
            next_parents = parents + (value,)
            for child in value.values():
                _walk(child, next_parents)
        elif isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
            for item in value:
                _walk(item, parents)

    for key in ("courses", "races", "items", "data"):
        if key in payload:
            _walk(payload[key], (payload,))

    # Deduplicate potential duplicates by object identity while preserving order
    seen_ids: set[int] = set()
    unique_results: list[tuple[Mapping[str, Any], tuple[Mapping[str, Any], ...]]] = []
    for course_payload, parents in results:
        obj_id = id(course_payload)
        if obj_id in seen_ids:
            continue
        seen_ids.add(obj_id)
        unique_results.append((course_payload, parents))

    return unique_results


def _upsert(
    ws: Worksheet, header_map: Mapping[str, int], row: Mapping[str, Any], keys: Sequence[str]
) -> None:
    key_values = {key: row.get(key) for key in keys}
    for key, value in key_values.items():
        if value in (None, ""):
            raise ValueError(f"Missing key value for {key}")

    target_row = None
    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        if all(
            str(ws.cell(row=row_idx, column=header_map[key]).value or "") == str(key_values[key])
            for key in keys
        ):
            target_row = row_idx
            break
    if target_row is None:
        target_row = max(2, max_row + 1)
    for header, value in row.items():
        if header not in header_map or value is None:
            continue
        ws.cell(row=target_row, column=header_map[header], value=value)


def _normalise_ticket_label(label: Any) -> str:
    if label in (None, ""):
        return ""
    text = str(label).strip()
    if not text:
        return ""
    if re.fullmatch(r"[A-Za-z0-9]{1,10}", text):
        return text.upper()
    return text


def _consume_selection(values: Any, horses: list[str]) -> None:
    if values in (None, ""):
        return
    if isinstance(values, Mapping):
        for key in ("selections", "horses", "selection", "horse", "num", "number", "id"):
            if key in values:
                _consume_selection(values.get(key), horses)
                return
        return
    if isinstance(values, Sequence) and not isinstance(values, (str, bytes, bytearray)):
        for item in values:
            _consume_selection(item, horses)
        return

    text = str(values).strip()
    if not text:
        return
    parts = re.split(r"[-,;/\\s]+", text)
    horses.extend(part for part in parts if part)


def _summarise_tickets(tickets: Any) -> str:
    if not isinstance(tickets, Iterable) or isinstance(tickets, (str, bytes, bytearray)):
        return ""
    summaries: list[str] = []
    for ticket in tickets:
        if not isinstance(ticket, Mapping):
            continue
        label = _first_non_empty(
            ticket.get("type"),
            ticket.get("bet_type"),
            ticket.get("label"),
            ticket.get("id"),
        )
        legs = ticket.get("legs")
        horses: list[str] = []
        if isinstance(legs, Iterable) and not isinstance(legs, (str, bytes)):
            for leg in legs:
                if isinstance(leg, Mapping):
                    _consume_selection(leg, horses)
                else:
                    _consume_selection(leg, horses)
        if not horses:
            for key in (
                "selections",
                "horses",
                "selection",
                "horse",
                "combination",
                "combo",
                "numbers",
            ):
                if key in ticket:
                    _consume_selection(ticket.get(key), horses)
        odds = _first_non_empty(ticket.get("odds"), ticket.get("rapport"), ticket.get("payout"))
        horses_joined = "-".join(str(horse) for horse in horses if str(horse))
        base = ""
        normalised_label = _normalise_ticket_label(label)
        if normalised_label:
            base = normalised_label
        if horses_joined:
            base = f"{base}:{horses_joined}" if base else horses_joined
        parts: list[str] = []
        if base:
            parts.append(base)
        if odds not in (None, ""):
            suffix = f"@{_coerce_number(odds)}"
            if parts:
                parts[-1] = f"{parts[-1]}{suffix}" if parts[-1] else suffix
            else:
                parts.append(suffix)
        segment = " ".join(parts)
        if segment:
            summaries.append(segment)
    return " | ".join(summaries)


def _extract_roi(payload: Mapping[str, Any]) -> float | None:
    candidates = [
        payload.get("roi"),
        payload.get("roi_est"),
    ]
    validation = payload.get("validation")
    if isinstance(validation, Mapping):
        candidates.extend(
            [
                validation.get("roi_global_est"),
                (validation.get("sp") or {}).get("roi_est")
                if isinstance(validation.get("sp"), Mapping)
                else None,
            ]
        )
    for candidate in candidates:
        try:
            return float(candidate)
        except (TypeError, ValueError):
            continue
    return None


def _status_h30(default: str) -> str:
    return default


def _status_h5(label: str) -> str:
    return label


def _jouable_flag(payload: Mapping[str, Any]) -> str:
    return "Non" if bool(payload.get("abstain")) else "Oui"


def _comment_h5(payload: Mapping[str, Any]) -> str:
    abstain = bool(payload.get("abstain"))
    reason = _first_non_empty(
        payload.get("abstain_reason"),
        payload.get("notes"),
        payload.get("message"),
    )
    roi = _extract_roi(payload)
    if abstain:
        return reason or ""
    if roi is not None:
        return f"ROI estimé {roi * 100:.0f}%"
    return reason or ""


def _collect_h30_entries(source: Path, status: str) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    if source.is_file():
        payloads = [(source, _load_json(source))]
    elif source.is_dir():
        payloads = [(path, _load_json(path)) for path in sorted(source.rglob("*.json"))]
    else:
        raise FileNotFoundError(source)
    for _path, payload in payloads:
        if not isinstance(payload, Mapping):
            continue
        courses = _extract_course_payloads(payload)
        if not courses:
            meta = _extract_common_meta(payload)
            if not (meta.get("date") and meta.get("reunion") and meta.get("course")):
                continue
            row = {
                "Date": meta.get("date"),
                "Réunion": meta.get("reunion"),
                "Course": meta.get("course"),
                "Hippodrome": _blank_if_missing(meta.get("hippodrome")),
                "Heure": _blank_if_missing(meta.get("start_time")),
                "Partants": _blank_if_missing(meta.get("partants")),
                "Discipline": _blank_if_missing(meta.get("discipline")),
                "Statut H-30": _status_h30(status),
                "Commentaires": None,
            }
            entries.append(row)
            continue

        for course_payload, parents in courses:
            meta = _extract_common_meta(course_payload, parents=parents)
            if not (meta.get("date") and meta.get("reunion") and meta.get("course")):
                continue
            row = {
                "Date": meta.get("date"),
                "Réunion": meta.get("reunion"),
                "Course": meta.get("course"),
                "Hippodrome": _blank_if_missing(meta.get("hippodrome")),
                "Heure": _blank_if_missing(meta.get("start_time")),
                "Partants": _blank_if_missing(meta.get("partants")),
                "Discipline": _blank_if_missing(meta.get("discipline")),
                "Statut H-30": _status_h30(status),
                "Commentaires": None,
            }
            entries.append(row)
    entries.sort(
        key=lambda row: (
            str(row.get("Date") or ""),
            str(row.get("Réunion") or ""),
            str(row.get("Course") or ""),
        )
    )
    return entries


def _load_h5_payload(source: Path) -> Mapping[str, Any]:
    if source.is_file():
        return _load_json(source)
    if source.is_dir():
        pattern_matches = sorted(source.glob("analysis_H5*.json"))
        for candidate in pattern_matches:
            if candidate.is_file():
                return _load_json(candidate)
        for name in ("analysis_H5.json", "analysis.json"):
            candidate = source / name
            if candidate.is_file():
                return _load_json(candidate)
    raise FileNotFoundError(source)


def _prepare_h5_row(payload: Mapping[str, Any], status_label: str) -> dict[str, Any]:
    meta = _extract_common_meta(payload)
    tickets = payload.get("tickets") if isinstance(payload, Mapping) else None
    summary = _summarise_tickets(tickets)
    row: dict[str, Any] = {
        "Date": meta.get("date"),
        "Réunion": meta.get("reunion"),
        "Course": meta.get("course"),
        "Hippodrome": _blank_if_missing(meta.get("hippodrome")),
        "Heure": _blank_if_missing(meta.get("start_time")),
        "Partants": _blank_if_missing(meta.get("partants")),
        "Discipline": _blank_if_missing(meta.get("discipline")),
        "Statut H-5": _status_h5(status_label),
        "Jouable H-5": _jouable_flag(payload),
    }
    row["Tickets H-5"] = summary or ""
    row["Commentaires"] = _comment_h5(payload) or ""
    return row


def main(argv: Sequence[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Met à jour l'onglet Planning de l'Excel")
    parser.add_argument(
        "--phase", required=True, type=_phase_argument, help="Phase à appliquer (H30/H5)"
    )
    parser.add_argument(
        "--input", "--in", dest="source", required=True, help="Fichier ou dossier source"
    )
    parser.add_argument("--excel", required=True, help="Chemin du classeur Excel")
    parser.add_argument("--sheet", default="Planning", help="Nom de l'onglet à mettre à jour")
    parser.add_argument(
        "--status-h30",
        default="Collecté",
        help="Statut par défaut renseigné pour les lignes H-30",
    )
    parser.add_argument(
        "--status-h5",
        default="Analysé",
        help="Statut renseigné pour les lignes H-5",
    )
    parser.add_argument("--rc", help=argparse.SUPPRESS)
    parser.add_argument("--drive-folder", help=argparse.SUPPRESS)
    args = parser.parse_args(argv)

    source = Path(args.source)
    excel_path = Path(args.excel)
    wb, _ = _load_workbook(excel_path)
    ws = _prepare_sheet(wb, args.sheet)
    header_map = _ensure_headers(ws, PLANNING_HEADERS)

    if args.phase == "H30":
        rows = _collect_h30_entries(source, status=args.status_h30)
        keys = ("Date", "Réunion", "Course")
        for row in rows:
            _upsert(ws, header_map, row, keys)
        details = ", ".join(filter(None, (_format_row_identifier(row) for row in rows)))
        suffix = f" ({details})" if details else ""
        message = f"{len(rows)} ligne(s) H-30 mises à jour{suffix}"
    else:
        payload = _load_h5_payload(source)
        row = _prepare_h5_row(payload, status_label=args.status_h5)
        keys = ("Date", "Réunion", "Course")
        if not all(row.get(key) for key in keys):
            raise ValueError("Impossible de déterminer Date/Réunion/Course pour la mise à jour H-5")
        _upsert(ws, header_map, row, keys)
        detail = _format_row_identifier(row)
        suffix = f" ({detail})" if detail else ""
        message = f"1 ligne H-5 mise à jour{suffix}"

    wb.save(excel_path)
    print(message)


if __name__ == "__main__":  # pragma: no cover
    main()
