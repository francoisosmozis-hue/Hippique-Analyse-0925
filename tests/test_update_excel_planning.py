
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src import update_excel_planning as planner


def _read_row(ws, row_idx: int) -> dict[str, object]:
    header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    return {header: ws.cell(row=row_idx, column=col).value for header, col in header_map.items()}


def test_cli_invalid_phase_raises_error():
    """Tests that the CLI exits if an invalid phase is provided."""
    with pytest.raises(SystemExit):
        planner.main(["--phase", "INVALID", "--input", "foo", "--excel", "bar"])


@pytest.mark.parametrize(
    "key, value, expected_key, expected_value",
    [
        ("nb_partants", 10, "partants", 10),
        ("runners_count", "11 partants", "partants", 11),
        ("start", "10h30", "start_time", "10:30"),
        ("heure_depart", "11:00", "start_time", "11:00"),
        ("specialite", "Steeple", "discipline", "Steeple"),
        ("r", "R5", "reunion", "R5"),
        ("c", "C6", "course", "C6"),
    ],
)
def test_extract_common_meta_aliases(key, value, expected_key, expected_value):
    """Tests that _extract_common_meta correctly uses aliased keys."""
    payload = {key: value}
    meta = planner._extract_common_meta(payload)
    assert meta[expected_key] == expected_value

    payload_nested = {"meta": {key: value}}
    meta_nested = planner._extract_common_meta(payload_nested)
    assert meta_nested[expected_key] == expected_value


@pytest.mark.parametrize(
    "tickets, expected_summary",
    [
        ([], ""),
        ([{"type": "SG", "selections": ["1"], "odds": 4.5}], "SG:1@4.5"),
        (
            [
                {"bet_type": "JUM", "horses": ["2", "4"], "rapport": 12.1},
                {"type": "TRIO", "combination": "5-6-7"},
            ],
            "JUM:2-4@12.1 | TRIO:5-6-7",
        ),
        ([{"id": "WEIRD", "legs": [{"selections": "8/9"}]}], "WEIRD:8-9"),
        ([{"type": "2S4", "numbers": "10,11,12"}], "2S4:10-11-12"),
    ],
)
def test_summarise_tickets_formats(tickets, expected_summary):
    """Tests that _summarise_tickets handles various ticket formats."""
    summary = planner._summarise_tickets(tickets)
    assert summary == expected_summary


def test_h30_updates_existing_row(tmp_path: Path):
    """Tests that running H30 phase again updates the existing row instead of creating a new one."""
    excel_path = tmp_path / "planning.xlsx"

    dir1 = tmp_path / "dir1"
    dir1.mkdir()
    payload1 = {
        "meta": {"date": "2024-11-17", "reunion": "R1", "course": "C1", "partants": 10}
    }
    (dir1 / "p1.json").write_text(json.dumps(payload1))
    planner.main(["--phase", "H30", "--input", str(dir1), "--excel", str(excel_path)])

    dir2 = tmp_path / "dir2"
    dir2.mkdir()
    payload2 = {
        "meta": {"date": "2024-11-17", "reunion": "R1", "course": "C1", "partants": 12, "discipline": "Trot"}
    }
    (dir2 / "p2.json").write_text(json.dumps(payload2))
    planner.main(["--phase", "H30", "--input", str(dir2), "--excel", str(excel_path), "--status-h30", "Mis à jour"])

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    assert ws.max_row == 2
    row = _read_row(ws, 2)
    assert row["Partants"] == 12
    assert row["Discipline"] == "Trot"
    assert row["Statut H-30"] == "Mis à jour"


def test_h30_populates_planning_sheet(tmp_path: Path) -> None:
    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    payload = {
        "meta": {
            "date": "2024-09-25",
            "reunion": "R1",
            "course": "C1",
            "hippodrome": "Paris-Vincennes",
            "start_time": "2024-09-25T12:05:00",
            "discipline": "Trot",
        },
        "runners": [{"id": 1}, {"id": 2}, {"id": 3}],
    }
    (meeting_dir / "snapshot_H-30.json").write_text(json.dumps(payload), encoding="utf-8")

    excel_path = tmp_path / "planning.xlsx"
    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)
    assert row["Date"] == "2024-09-25"
    assert row["Heure"] == "12:05"
    assert row["Partants"] == 3
    assert row["Discipline"] == "Trot"
    assert row["Statut H-30"] == "Collecté"
    assert row.get("Statut H-5") in (None, "")
    assert row.get("Jouable H-5") in (None, "")


def test_h5_updates_status_and_tickets(tmp_path: Path) -> None:
    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    meeting_payload = {
        "meta": {
            "date": "2024-09-25",
            "reunion": "R1",
            "course": "C1",
            "hippodrome": "Paris-Vincennes",
            "discipline": "Trot",
        },
        "runners": [{"id": 1}, {"id": 2}],
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(meeting_payload), encoding="utf-8")
    excel_path = tmp_path / "planning.xlsx"
    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    rc_dir = tmp_path / "R1C1"
    rc_dir.mkdir()
    analysis = {
        "meta": {
            "date": "2024-09-25",
            "reunion": "R1",
            "course": "C1",
            "hippodrome": "Paris-Vincennes",
            "start_time": "12:10",
            "discipline": "Trot",
        },
        "validation": {"roi_global_est": 0.31},
        "abstain": False,
        "tickets": [
            {
                "id": "SP1",
                "type": "SP",
                "legs": [{"horse": "6"}],
                "stake": 2.0,
                "odds": 3.4,
            }
        ],
    }
    (rc_dir / "analysis_H5_R1C1.json").write_text(json.dumps(analysis), encoding="utf-8")

    planner.main(
        [
            "--phase",
            "H5",
            "--input",
            str(rc_dir),
            "--excel",
            str(excel_path),
            "--rc",
            "R1C1",
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)
    assert row["Statut H-30"] == "Collecté"
    assert row["Statut H-5"] == "Analysé"
    assert row["Jouable H-5"] == "Oui"
    assert row["Commentaires"] == "ROI estimé 31%"
    assert row["Tickets H-5"] == "SP:6@3.4"
    assert row["Heure"] == "12:10"


def test_h5_abstention_uses_reason(tmp_path: Path) -> None:
    excel_path = tmp_path / "planning.xlsx"
    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    base_payload = {
        "meta": {
            "date": "2024-09-25",
            "reunion": "R1",
            "course": "C1",
        },
        "runners": [{"id": 1}],
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(base_payload), encoding="utf-8")
    planner.main([
        "--phase",
        "H30",
        "--input",
        str(meeting_dir),
        "--excel",
        str(excel_path),
    ])

    rc_dir = tmp_path / "R1C1"
    rc_dir.mkdir()
    abstain_payload = {
        "meta": {
            "date": "2024-09-25",
            "reunion": "R1",
            "course": "C1",
        },
        "abstain": True,
        "notes": "ROI global < 0.20",
    }
    (rc_dir / "analysis.json").write_text(json.dumps(abstain_payload), encoding="utf-8")

    planner.main([
        "--phase",
        "H5",
        "--input",
        str(rc_dir / "analysis.json"),
        "--excel",
        str(excel_path),
        "--rc",
        "R1C1",
    ])

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)
    assert row["Statut H-5"] == "Analysé"
    assert row["Jouable H-5"] == "Non"
    assert row["Commentaires"] == "ROI global < 0.20"
    assert row.get("Tickets H-5") in (None, "")


def test_h5_preserves_existing_metadata_when_missing(tmp_path: Path) -> None:
    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    base_payload = {
        "meta": {
            "date": "2024-09-26",
            "reunion": "R2",
            "course": "C3",
            "hippodrome": "Lyon-Parilly",
            "start_time": "2024-09-26T14:15:00",
        },
        "runners": [{"id": 1}, {"id": 2}],
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(base_payload), encoding="utf-8")

    excel_path = tmp_path / "planning.xlsx"
    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    analysis_dir = tmp_path / "R2C3"
    analysis_dir.mkdir()
    minimal_h5 = {
        "meta": {
            "date": "2024-09-26",
            "reunion": "R2",
            "course": "C3",
        },
        "abstain": True,
        "abstain_reason": "Champ trop ouvert",
    }
    (analysis_dir / "analysis_H5.json").write_text(json.dumps(minimal_h5), encoding="utf-8")

    planner.main(
        [
            "--phase",
            "H5",
            "--input",
            str(analysis_dir),
            "--excel",
            str(excel_path),
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)
    assert row["Hippodrome"] == "Lyon-Parilly"
    assert row["Heure"] == "14:15"
    assert row["Statut H-5"] == "Analysé"
    assert row["Jouable H-5"] == "Non"


def test_h5_converts_timezone_when_tz_env_set(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    base_payload = {
        "meta": {
            "date": "2024-10-01",
            "reunion": "R3",
            "course": "C4",
            "hippodrome": "Chantilly",
        },
        "runners": [{"id": 1}, {"id": 2}],
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(base_payload), encoding="utf-8")

    excel_path = tmp_path / "planning.xlsx"
    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    monkeypatch.setenv("TZ", "Europe/Paris")
    planner._env_timezone.cache_clear()

    rc_dir = tmp_path / "R3C4"
    rc_dir.mkdir()
    analysis_payload = {
        "meta": {
            "date": "2024-10-01",
            "reunion": "R3",
            "course": "C4",
            "start_time": "2024-10-01T12:05:00+00:00",
        },
        "abstain": False,
    }
    (rc_dir / "analysis_H5.json").write_text(json.dumps(analysis_payload), encoding="utf-8")

    planner.main(
        [
            "--phase",
            "H5",
            "--input",
            str(rc_dir),
            "--excel",
            str(excel_path),
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)
    assert row["Heure"] == "14:05"

    planner._env_timezone.cache_clear()


def test_custom_status_h5_flag(tmp_path: Path) -> None:
    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    payload = {
        "meta": {
            "date": "2024-09-25",
            "reunion": "R1",
            "course": "C1",
        },
        "runners": [{"id": 1}],
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(payload), encoding="utf-8")

    excel_path = tmp_path / "planning.xlsx"
    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    analysis_dir = tmp_path / "R1C1"
    analysis_dir.mkdir()
    analysis_payload = {
        "meta": {
            "date": "2024-09-25",
            "reunion": "R1",
            "course": "C1",
        },
        "tickets": [],
    }
    (analysis_dir / "analysis.json").write_text(json.dumps(analysis_payload), encoding="utf-8")

    planner.main(
        [
            "--phase",
            "H5",
            "--input",
            str(analysis_dir),
            "--excel",
            str(excel_path),
            "--status-h5",
            "Validé",
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)
    assert row["Statut H-5"] == "Validé"


def test_custom_status_h30_and_h5(tmp_path: Path) -> None:
    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    payload = {
        "meta": {
            "date": "2024-09-30",
            "reunion": "R2",
            "course": "C3",
        },
        "runners": [{"id": 1}],
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(payload), encoding="utf-8")

    excel_path = tmp_path / "planning.xlsx"
    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
            "--status-h30",
            "Importé",
        ]
    )

    analysis_dir = tmp_path / "R2C3"
    analysis_dir.mkdir()
    analysis_payload = {
        "meta": {
            "date": "2024-09-30",
            "reunion": "R2",
            "course": "C3",
        },
        "abstain": False,
        "tickets": [],
    }
    (analysis_dir / "analysis_H5.json").write_text(json.dumps(analysis_payload), encoding="utf-8")

    planner.main(
        [
            "--phase",
            "H5",
            "--input",
            str(analysis_dir),
            "--excel",
            str(excel_path),
            "--status-h5",
            "Validé",
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)
    assert row["Statut H-30"] == "Importé"
    assert row["Statut H-5"] == "Validé"


def test_h30_preserves_existing_comment(tmp_path: Path) -> None:
    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    payload = {
        "meta": {
            "date": "2024-09-27",
            "reunion": "R1",
            "course": "C2",
            "hippodrome": "Enghien",
        },
        "runners": [{"id": 1}, {"id": 2}],
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(payload), encoding="utf-8")

    excel_path = tmp_path / "planning.xlsx"
    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    ws.cell(row=2, column=header_map["Commentaires"]).value = "Note à conserver"
    wb.save(excel_path)
    wb.close()

    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)
    assert row["Commentaires"] == "Note à conserver"


def test_h30_handles_nested_course_payload(tmp_path: Path) -> None:
    meeting_dir = tmp_path / "snapshots"
    meeting_dir.mkdir()
    payload = {
        "courses": [
            {
                "meta": {
                    "date": "2024-09-25",
                    "reunion": "R4",
                    "course": "C5",
                    "hippodrome": "Vincennes",
                    "start_time": "13h05",
                    "partants": 14,
                },
                "discipline": "Plat",
            }
        ]
    }
    (meeting_dir / "meeting.json").write_text(json.dumps(payload), encoding="utf-8")
    excel_path = tmp_path / "planning.xlsx"

    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)

    assert row["Réunion"] == "R4"
    assert row["Course"] == "C5"
    assert row["Hippodrome"] == "Vincennes"
    assert row["Heure"] == "13:05"
    assert row["Partants"] == 14
    assert row["Discipline"] == "Plat"


def test_h30_infers_rc_fields(tmp_path: Path) -> None:
    meeting_dir = tmp_path / "snapshots"
    meeting_dir.mkdir()
    payload = {
        "meta": {
            "date": "2024-10-01",
            "rc": "R4C5",
            "hippodrome": "Laval",
        },
        "runners": [{"id": 1}, {"id": 2}],
    }
    (meeting_dir / "meeting.json").write_text(json.dumps(payload), encoding="utf-8")
    excel_path = tmp_path / "planning.xlsx"

    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)

    assert row["Réunion"] == "R4"
    assert row["Course"] == "C5"
    assert row["Hippodrome"] == "Laval"


def test_h5_infers_rc_fields(tmp_path: Path) -> None:
    excel_path = tmp_path / "planning.xlsx"
    analysis_file = tmp_path / "analysis.json"
    payload = {
        "meta": {
            "date": "2024-10-02",
            "rc": "r7 c3",
            "start_time": "13h30",
        },
        "tickets": [],
        "abstain": False,
    }
    analysis_file.write_text(json.dumps(payload), encoding="utf-8")

    planner.main(
        [
            "--phase",
            "H5",
            "--input",
            str(analysis_file),
            "--excel",
            str(excel_path),
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)

    assert row["Réunion"] == "R7"
    assert row["Course"] == "C3"
    assert row["Heure"] == "13:30"


def test_h5_compact_ticket_summary(tmp_path: Path) -> None:
    meeting_dir = tmp_path / "snapshots"
    meeting_dir.mkdir()
    payload = {
        "courses": [
            {
                "meta": {
                    "date": "2024-09-25",
                    "reunion": "R4",
                    "course": "C5",
                    "hippodrome": "Vincennes",
                },
            }
        ]
    }
    (meeting_dir / "meeting.json").write_text(json.dumps(payload), encoding="utf-8")
    excel_path = tmp_path / "planning.xlsx"
    planner.main(
        [
            "--phase",
            "H30",
            "--input",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )

    analysis_dir = tmp_path / "R4C5"
    analysis_dir.mkdir()
    analysis_payload = {
        "meta": {
            "date": "2024-09-25",
            "reunion": "R4",
            "course": "C5",
            "hippodrome": "Vincennes",
            "start_time": "13h05",
            "discipline": "Plat",
            "partants": 14,
        },
        "abstain": False,
        "roi": 0.12,
        "tickets": [
            {
                "type": "SP",
                "legs": [{"selections": [3, 5]}],
                "odds": 2.0,
            },
            {
                "type": "CPL",
                "legs": [{"horses": [1, 3]}],
                "rapport": 1.5,
            },
        ],
    }
    (analysis_dir / "analysis_H5.json").write_text(json.dumps(analysis_payload), encoding="utf-8")

    planner.main(
        [
            "--phase",
            "H5",
            "--input",
            str(analysis_dir),
            "--excel",
            str(excel_path),
        ]
    )

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)

    assert row["Statut H-5"] == "Analysé"
    assert row["Jouable H-5"] == "Oui"
    assert row["Tickets H-5"] == "SP:3-5@2 | CPL:1-3@1.5"
    assert row["Commentaires"] == "ROI estimé 12%"

def test_format_time_respects_input_timezone(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("TZ", raising=False)
    planner._env_timezone.cache_clear()
    assert planner._format_time("2024-09-25T13:05:00+02:00") == "13:05"


def test_format_time_uses_env_timezone(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("TZ", "Europe/Paris")
    planner._env_timezone.cache_clear()
    assert planner._format_time("2024-09-25T13:05:00+00:00") == "15:05"
    monkeypatch.delenv("TZ", raising=False)
    planner._env_timezone.cache_clear()


@pytest.mark.parametrize(
    "value, expected",
    [
        ("13h05", "13:05"),
        ("7h", "07:00"),
        ("09 heures 30", "09:30"),
        ("18.45", "18:45"),
    ],
)
def test_format_time_handles_textual_hours(value: str, expected: str) -> None:
    """French textual hour formats should be normalised to HH:MM."""

    assert planner._format_time(value) == expected


def test_cli_alias_and_dash_phase(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    """The CLI should accept ``--in`` and dashed phase labels while printing a summary."""

    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    meeting_payload = {
        "meta": {
            "date": "2025-09-26",
            "reunion": "R4",
            "course": "C5",
            "hippodrome": "Vincennes",
            "start_time": "13h15",
            "discipline": "Trot",
        },
        "runners": [{"id": 1}, {"id": 2}, {"id": 3}, {"id": 4}],
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(meeting_payload), encoding="utf-8")

    excel_path = tmp_path / "planning.xlsx"

    planner.main(
        [
            "--phase",
            "H-30",
            "--in",
            str(meeting_dir),
            "--excel",
            str(excel_path),
        ]
    )
    out_h30 = capsys.readouterr().out
    assert "1 ligne(s) H-30 mises à jour" in out_h30
    assert "R4C5" in out_h30
    assert "2025-09-26" in out_h30

    analysis_dir = tmp_path / "analysis"
    analysis_dir.mkdir()
    analysis_payload = {
        "meta": {
            "date": "2025-09-26",
            "reunion": "R4",
            "course": "C5",
            "hippodrome": "Vincennes",
            "start_time": "13:15",
            "discipline": "Trot",
        },
        "tickets": [
            {
                "type": "SP",
                "legs": [{"selections": [3, 5]}],
                "odds": 2.0,
            }
        ],
        "abstain": False,
        "roi": 0.4,
    }
    (analysis_dir / "analysis_H5.json").write_text(json.dumps(analysis_payload), encoding="utf-8")

    planner.main(
        [
            "--phase",
            "H5",
            "--in",
            str(analysis_dir),
            "--excel",
            str(excel_path),
        ]
    )
    out_h5 = capsys.readouterr().out
    assert "1 ligne H-5 mise à jour" in out_h5
    assert "R4C5" in out_h5
    assert "2025-09-26" in out_h5

    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    row = _read_row(ws, 2)
    assert row["Date"] == "2025-09-26"
    assert row["Réunion"] == "R4"
    assert row["Course"] == "C5"
    assert row["Heure"] == "13:15"
    assert row["Partants"] == 4
    assert row["Statut H-30"] == "Collecté"
    assert row["Statut H-5"] == "Analysé"
    assert row["Jouable H-5"] == "Oui"
    assert row["Tickets H-5"] == "SP:3-5@2"
    assert row["Commentaires"] == "ROI estimé 40%"
