from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook


def test_update_excel_records_observed_roi(tmp_path: Path) -> None:
    excel_path = tmp_path / "modele.xlsx"

    tickets_data = {
        "meta": {
            "rc": "R1C1",
            "hippodrome": "Test",
            "date": "2025-09-10",
            "discipline": "trot",
            "model": "GPI",
        },
        "tickets": [
            {
                "id": "1",
                "stake": 2.0,
                "odds": 5.0,
                "p": 0.3,
                "gain_reel": 10.0,
                "roi_reel": 4.0,
                "result": 1,
                "brier": 0.49,
            },
            {
                "id": "2",
                "stake": 1.0,
                "odds": 3.0,
                "p": 0.2,
                "gain_reel": 0.0,
                "roi_reel": -1.0,
                "result": 0,
                "brier": 0.04,
            },
        ],
        "roi_reel": (10.0 - 3.0) / 3.0,
        "roi_reel_moyen": (4.0 + -1.0) / 2,
        "result_moyen": 0.5,
        "brier_total": 0.53,
        "brier_moyen": 0.265,
        "ev": {
            "sp": 1.5,
            "global": 2.0,
            "roi_sp": 0.5,
            "roi_global": 0.7,
            "risk_of_ruin": 0.1,
            "clv_moyen": 0.2,
            "variance": 0.05,
            "combined_expected_payout": 9.0,
        },
    }
    tickets_path = tmp_path / "tickets.json"
    tickets_path.write_text(json.dumps(tickets_data), encoding="utf-8")

    arrivee_path = tmp_path / "arrivee_officielle.json"
    arrivee_path.write_text(
        json.dumps({"rc": "R1C1", "date": "2025-09-10"}),
        encoding="utf-8",
    )

    cmd = [
        sys.executable,
        "-m",
        "hippique_orchestrator.scripts.update_excel_with_results",
        "--excel",
        str(excel_path),
        "--arrivee",
        str(arrivee_path),
        "--tickets",
        str(tickets_path),
    ]
    res = subprocess.run(cmd, check=False, capture_output=True, text=True)
    assert res.returncode == 0, res.stderr
    lines = [line for line in res.stdout.splitlines() if line.startswith("Suivi:")]
    assert lines, res.stdout
    payload_line = lines[-1].split("Suivi:", 1)[1].strip()
    printed = json.loads(payload_line)
    assert printed["R/C"] == "R1C1"
    assert printed["ROI_reel"] == pytest.approx((10.0 - 3.0) / 3.0, rel=1e-6)

    wb = load_workbook(excel_path)
    assert "Sheet" not in wb.sheetnames

    ws_observe = wb["ROI Observé"]
    assert ws_observe.max_row >= 2
    header_map_observe = {
        ws_observe.cell(row=1, column=col).value: ws_observe.cell(row=2, column=col).value
        for col in range(1, ws_observe.max_column + 1)
        if ws_observe.cell(row=1, column=col).value
    }
    assert header_map_observe["R/C"] == "R1C1"
    assert header_map_observe["mises"] == pytest.approx(3.0)
    assert header_map_observe["gains"] == pytest.approx(10.0)
    assert header_map_observe["ROI_reel"] == pytest.approx(tickets_data["roi_reel"])
    assert header_map_observe["ROI_reel_moyen"] == pytest.approx(tickets_data["roi_reel_moyen"])

    ws_prevision = wb["ROI Prévisionnel"]
    assert ws_prevision.max_row >= 2
    header_map_prevision = {
        ws_prevision.cell(row=1, column=col).value: ws_prevision.cell(row=2, column=col).value
        for col in range(1, ws_prevision.max_column + 1)
        if ws_prevision.cell(row=1, column=col).value
    }
    assert header_map_prevision["ROI_global"] == pytest.approx(
        tickets_data["ev"]["roi_global"]
    )

    ws_suivi = wb["Suivi"]
    header_map_suivi = {
        ws_suivi.cell(row=1, column=col).value: ws_suivi.cell(row=2, column=col).value
        for col in range(1, ws_suivi.max_column + 1)
        if ws_suivi.cell(row=1, column=col).value
    }
    assert header_map_suivi["ROI_reel"] == pytest.approx((10.0 - 3.0) / 3.0)
    assert header_map_suivi["ROI_estime"] == pytest.approx(tickets_data["ev"]["roi_global"])
