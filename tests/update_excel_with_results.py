#!/usr/bin/env python3
"""Update ROI worksheets in the tracking Excel workbook.

This script is meant to be called from both the *pipeline* stage
(`pipeline_run.py`) and the *post-course* workflow (`post_course.py`). It
combines the information contained in the exported ``tickets`` artefact with
the official arrival JSON to append or update rows in two dedicated Excel
worksheets:

* ``ROI Prévisionnel`` – contains the simulated EV/ROI metrics prior to the
  race.
* ``ROI Observé`` – stores the realised ROI once the official result is
  known.

If the workbook does not exist yet it will be created on the fly. Rows are
deduplicated based on the course identifier (``R/C`` column): running the
script twice for the same race simply refreshes the existing line.
"""

from __future__ import annotations

import argparse
import json
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

JsonDict = dict[str, Any]


PREVISION_HEADERS = [
    "R/C",
    "hippodrome",
    "date",
    "discipline",
    "mises",
    "EV_sp",
    "EV_global",
    "ROI_sp",
    "ROI_global",
    "risk_of_ruin",
    "clv_moyen",
    "variance",
    "payout_attendu",
    "model",
]


OBSERVED_HEADERS = [
    "R/C",
    "hippodrome",
    "date",
    "discipline",
    "mises",
    "gains",
    "ROI_reel",
    "result_moyen",
    "ROI_reel_moyen",
    "Brier_total",
    "Brier_moyen",
    "EV_total",
    "EV_ecart",
    "model",
]


def _load_json(path: str | Path) -> JsonDict:
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def _load_workbook(path: Path) -> tuple[Workbook, bool]:
    if path.exists():
        return load_workbook(path), False
    return Workbook(), True


def _get_sheet(wb: Workbook, title: str) -> Worksheet:
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title=title)


def _ensure_header_map(ws: Worksheet, headers: Iterable[str]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    max_column = ws.max_column
    # ``openpyxl`` initialises empty worksheets with a single ``None`` cell.
    blank_sheet = (
        ws.max_row <= 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value in (None, "")
    )
    if not blank_sheet:
        for col in range(1, max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value not in (None, ""):
                header_map[str(value)] = col
    next_col = 0 if blank_sheet else max_column
    for header in headers:
        if header not in header_map:
            next_col += 1
            ws.cell(row=1, column=next_col, value=header)
            header_map[header] = next_col
    return header_map


def _upsert_row(
    ws: Worksheet,
    headers: Iterable[str],
    values: dict[str, Any],
    *,
    key_header: str = "R/C",
) -> None:
    header_map = _ensure_header_map(ws, headers)
    key_value = values.get(key_header)
    if key_value in (None, ""):
        return
    key_col = header_map[key_header]
    key_str = str(key_value)
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=key_col).value
        if cell_value is None:
            continue
        if str(cell_value) == key_str:
            target_row = row_idx
            break
    if target_row is None:
        target_row = max(ws.max_row + 1, 2)
    for header in headers:
        col_idx = header_map[header]
        ws.cell(row=target_row, column=col_idx, value=values.get(header))


def _merge_meta(arrivee: JsonDict, tickets: JsonDict) -> JsonDict:
    meta: JsonDict = {}
    tickets_meta = tickets.get("meta")
    if isinstance(tickets_meta, dict):
        meta.update(tickets_meta)
    arrivee_meta = arrivee.get("meta")
    if isinstance(arrivee_meta, dict):
        for key, value in arrivee_meta.items():
            meta.setdefault(key, value)
    for key in ("rc", "hippodrome", "date", "discipline", "model", "MODEL"):
        if not meta.get(key):
            value = arrivee.get(key)
            if value is None and isinstance(arrivee_meta, dict):
                value = arrivee_meta.get(key)
            if value is None and isinstance(tickets, dict):
                value = tickets.get(key)
            if value is not None:
                meta[key] = value
    if "model" not in meta and "MODEL" in meta:
        meta["model"] = meta["MODEL"]
    return meta


def _compute_ticket_metrics(
    tickets: Iterable[JsonDict],
) -> tuple[float, float, float, float, float, float, float]:
    total_stake = 0.0
    total_gain = 0.0
    total_ev = 0.0
    total_diff_ev = 0.0
    total_result = 0.0
    total_roi_ticket = 0.0
    total_brier = 0.0
    count = 0

    for ticket in tickets:
        stake = float(ticket.get("stake", 0.0) or 0.0)
        total_stake += stake
        gain = float(ticket.get("gain_reel", 0.0) or 0.0)
        total_gain += gain

        result = ticket.get("result")
        if result is not None:
            total_result += float(result)

        roi_ticket = ticket.get("roi_reel")
        if roi_ticket is not None:
            total_roi_ticket += float(roi_ticket)

        brier = ticket.get("brier")
        if brier is not None:
            total_brier += float(brier)

        ev_value = ticket.get("ev")
        if ev_value is None:
            ev_value = ticket.get("ev_ticket")
        if ev_value is None and {"p", "odds"}.issubset(ticket):
            try:
                p = float(ticket.get("p", 0.0) or 0.0)
                odds = float(ticket.get("odds", 0.0) or 0.0)
            except (TypeError, ValueError):
                ev_value = None
            else:
                ev_value = stake * (p * (odds - 1.0) - (1.0 - p))
        if ev_value is not None:
            ev_float = float(ev_value)
            total_ev += ev_float
            total_diff_ev += gain - ev_float

        count += 1

    mean_result = total_result / count if count else 0.0
    mean_roi_ticket = total_roi_ticket / count if count else 0.0
    mean_brier = total_brier / count if count else 0.0
    return (
        total_stake,
        total_gain,
        total_ev,
        total_diff_ev,
        mean_result,
        mean_roi_ticket,
        mean_brier,
    )


def _update_previsionnel_sheet(
    wb: Workbook,
    sheet_name: str,
    meta: JsonDict,
    tickets: JsonDict,
) -> None:
    rc = meta.get("rc")
    if not rc:
        return
    ws = _get_sheet(wb, sheet_name)
    total_stake, _, _, _, _, _, _ = _compute_ticket_metrics(tickets.get("tickets", []))
    ev_data = tickets.get("ev", {}) if isinstance(tickets, dict) else {}
    row = {
        "R/C": rc,
        "hippodrome": meta.get("hippodrome", ""),
        "date": meta.get("date", ""),
        "discipline": meta.get("discipline", ""),
        "mises": round(total_stake, 2) if total_stake else 0.0,
        "EV_sp": ev_data.get("sp"),
        "EV_global": ev_data.get("global"),
        "ROI_sp": ev_data.get("roi_sp"),
        "ROI_global": ev_data.get("roi_global"),
        "risk_of_ruin": ev_data.get("risk_of_ruin"),
        "clv_moyen": ev_data.get("clv_moyen"),
        "variance": ev_data.get("variance"),
        "payout_attendu": ev_data.get("combined_expected_payout"),
        "model": meta.get("model") or meta.get("MODEL", ""),
    }
    _upsert_row(ws, PREVISION_HEADERS, row)


def _update_observe_sheet(
    wb: Workbook,
    sheet_name: str,
    meta: JsonDict,
    tickets: JsonDict,
) -> None:
    rc = meta.get("rc")
    if not rc:
        return
    tickets_list = tickets.get("tickets", []) if isinstance(tickets, dict) else []
    (
        total_stake,
        total_gain,
        total_ev,
        total_diff_ev,
        mean_result,
        mean_roi_ticket,
        mean_brier,
    ) = _compute_ticket_metrics(tickets_list)

    # Determine whether we have observed data to store.
    roi_observed = tickets.get("roi_reel")
    if roi_observed is None and total_stake:
        roi_observed = (total_gain - total_stake) / total_stake if total_stake else 0.0
    has_observed_metrics = any(
        (
            roi_observed,
            total_gain,
            any("gain_reel" in t for t in tickets_list),
        )
    )
    if not has_observed_metrics:
        return

    row = {
        "R/C": rc,
        "hippodrome": meta.get("hippodrome", ""),
        "date": meta.get("date", ""),
        "discipline": meta.get("discipline", ""),
        "mises": round(total_stake, 2) if total_stake else 0.0,
        "gains": round(total_gain, 2) if total_gain else 0.0,
        "ROI_reel": roi_observed,
        "result_moyen": tickets.get("result_moyen", mean_result),
        "ROI_reel_moyen": tickets.get("roi_reel_moyen", mean_roi_ticket),
        "Brier_total": tickets.get("brier_total", mean_brier * len(tickets_list)),
        "Brier_moyen": tickets.get("brier_moyen", mean_brier),
        "EV_total": total_ev,
        "EV_ecart": total_diff_ev,
        "model": meta.get("model") or meta.get("MODEL", ""),
    }
    ws = _get_sheet(wb, sheet_name)
    _upsert_row(ws, OBSERVED_HEADERS, row)


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Mettre à jour le ROI dans l'Excel de suivi")
    parser.add_argument(
        "--excel",
        default="modele_suivi_courses_hippiques.xlsx",
        help="Chemin du classeur Excel à mettre à jour",
    )
    parser.add_argument("--arrivee", required=True, help="JSON de l'arrivée officielle")
    parser.add_argument("--tickets", required=True, help="JSON des tickets (p_finale.json)")
    parser.add_argument(
        "--sheet-prevision",
        default="ROI Prévisionnel",
        help="Nom de la feuille pour les projections",
    )
    parser.add_argument(
        "--sheet-observe",
        default="ROI Observé",
        help="Nom de la feuille pour le ROI observé",
    )
    args = parser.parse_args(argv)

    arrivee_path = Path(args.arrivee)
    tickets_path = Path(args.tickets)
    excel_path = Path(args.excel)

    arrivee_data = _load_json(arrivee_path)
    tickets_data = _load_json(tickets_path)
    meta = _merge_meta(arrivee_data, tickets_data)

    wb, created = _load_workbook(excel_path)
    if created and "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        default_ws = wb["Sheet"]
        if default_ws.max_row <= 1 and default_ws.cell(row=1, column=1).value in (None, ""):
            wb.remove(default_ws)

    _update_previsionnel_sheet(wb, args.sheet_prevision, meta, tickets_data)
    _update_observe_sheet(wb, args.sheet_observe, meta, tickets_data)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    main()
