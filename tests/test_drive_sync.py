import base64
import json
import sys
from pathlib import Path
from typing import Any
from unittest.mock import ANY, MagicMock, patch

import pytest
from google.auth.exceptions import DefaultCredentialsError
from openpyxl import load_workbook

pytest.importorskip("google.cloud.storage")

from scripts import drive_sync


def test_upload_file_uses_bucket_env(tmp_path, monkeypatch):
    path = tmp_path / "example.txt"
    path.write_text("data", encoding="utf-8")
    monkeypatch.setenv("GCS_BUCKET", "my-bucket")
    client = MagicMock()
    bucket = MagicMock()
    blob = MagicMock()
    bucket.blob.return_value = blob
    client.bucket.return_value = bucket

    drive_sync.upload_file(path, folder_id="prefix", service=client)

    client.bucket.assert_called_once_with("my-bucket")
    bucket.blob.assert_called_once_with("prefix/example.txt")
    blob.upload_from_filename.assert_called_once_with(str(path))


def test_upload_file_uses_env_prefix(tmp_path, monkeypatch):
    path = tmp_path / "env.txt"
    path.write_text("ok", encoding="utf-8")
    monkeypatch.setenv("GCS_BUCKET", "bucket")
    monkeypatch.setenv("GCS_PREFIX", "base/prefix")
    client = MagicMock()
    bucket = MagicMock()
    blob = MagicMock()
    bucket.blob.return_value = blob
    client.bucket.return_value = bucket

    drive_sync.upload_file(path, service=client)

    bucket.blob.assert_called_once_with("base/prefix/env.txt")


def test_upload_file_missing_bucket(tmp_path, monkeypatch):
    path = tmp_path / "missing.txt"
    path.write_text("x", encoding="utf-8")
    monkeypatch.delenv("GCS_BUCKET", raising=False)
    with pytest.raises(EnvironmentError):
        drive_sync.upload_file(path, service=MagicMock())


def test_download_file_calls_blob(tmp_path, monkeypatch):
    dest = tmp_path / "out" / "file.txt"
    monkeypatch.setenv("GCS_BUCKET", "bucket")
    client = MagicMock()
    bucket = MagicMock()
    blob = MagicMock()
    bucket.blob.return_value = blob
    client.bucket.return_value = bucket

    result = drive_sync.download_file("object/name.json", dest, service=client)

    assert result == dest
    assert dest.parent.exists()
    bucket.blob.assert_called_once_with("object/name.json")
    blob.download_to_filename.assert_called_once_with(str(dest))


def test_push_tree_uploads_all_files(tmp_path, monkeypatch):
    (tmp_path / "sub").mkdir()
    (tmp_path / "root.txt").write_text("r", encoding="utf-8")
    (tmp_path / "sub" / "child.txt").write_text("c", encoding="utf-8")

    monkeypatch.setenv("GCS_BUCKET", "bucket")
    client = MagicMock()
    bucket = MagicMock()
    created: dict[str, MagicMock] = {}

    def _blob(name: str):
        created[name] = MagicMock()
        return created[name]

    bucket.blob.side_effect = _blob
    client.bucket.return_value = bucket

    drive_sync.push_tree(tmp_path, folder_id="prefix", service=client)

    assert set(created) == {"prefix/root.txt", "prefix/sub/child.txt"}
    for blob in created.values():
        blob.upload_from_filename.assert_called_once()


def test_main_honours_env_prefix(monkeypatch):
    monkeypatch.setenv("GCS_BUCKET", "bucket")
    monkeypatch.setenv("GCS_PREFIX", "env/prefix")
    monkeypatch.setattr(drive_sync, "is_gcs_enabled", lambda: True)
    client = MagicMock(name="client")
    monkeypatch.setattr(drive_sync, "_build_service", lambda *a, **k: client)
    push_mock = MagicMock()
    monkeypatch.setattr(drive_sync, "push_tree", push_mock)
    monkeypatch.setattr(drive_sync, "upload_file", MagicMock())
    monkeypatch.setattr(drive_sync, "download_file", MagicMock())
    monkeypatch.setattr(drive_sync, "_iter_uploads", lambda patterns: [])
    monkeypatch.setattr(sys, "argv", ["drive_sync.py", "--push", "data"])

    result = drive_sync.main()

    assert result == 0
    push_mock.assert_called_once_with(
        "data", folder_id="env/prefix", bucket="bucket", service=client
    )


def _write_sample_payload(tmp_path: Path) -> tuple[Path, Path, Path, Path]:
    tickets_path = tmp_path / "tickets.json"
    arrivee_path = tmp_path / "arrivee_officielle.json"
    outdir = tmp_path / "artefacts"
    excel_path = tmp_path / "excel" / "roi.xlsx"

    tickets = {
        "meta": {
            "rc": "R1C1",
            "hippodrome": "Test",
            "date": "2024-09-25",
            "discipline": "trot",
            "model": "GPI",
        },
        "tickets": [
            {"id": "1", "stake": 2.0, "odds": 4.0, "p": 0.3},
            {"id": "3", "stake": 1.0, "odds": 2.5, "p": 0.2},
        ],
    }
    arrivee = {"rc": "R1C1", "result": ["1", "5", "7"]}
    tickets_path.write_text(json.dumps(tickets), encoding="utf-8")
    arrivee_path.write_text(json.dumps(arrivee), encoding="utf-8")
    return tickets_path, arrivee_path, outdir, excel_path


def test_main_skips_when_credentials_missing(monkeypatch, capsys, tmp_path):
    tickets_path, arrivee_path, outdir, excel_path = _write_sample_payload(tmp_path)

    monkeypatch.setattr(drive_sync, "is_gcs_enabled", lambda: True)
    monkeypatch.setattr(
        drive_sync,
        "push_tree",
        MagicMock(side_effect=AssertionError("push_tree should not run")),
    )
    monkeypatch.setattr(
        drive_sync,
        "upload_file",
        MagicMock(side_effect=AssertionError("upload_file should not run")),
    )
    monkeypatch.setattr(
        drive_sync,
        "download_file",
        MagicMock(side_effect=AssertionError("download_file should not run")),
    )

    def _raise_missing(*_args, **_kwargs):
        raise DefaultCredentialsError("missing")

    monkeypatch.setattr(drive_sync, "_build_service", _raise_missing)
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "drive_sync.py",
            "--arrivee",
            str(arrivee_path),
            "--tickets",
            str(tickets_path),
            "--outdir",
            str(outdir),
            "--excel",
            str(excel_path),
        ],
    )

    result = drive_sync.main()

    captured = capsys.readouterr()
    assert result == 0
    assert "[drive_sync] ROI non historisé (Drive off)" in captured.out
    assert (outdir / "arrivee.json").exists()
    assert (outdir / "ligne_resultats.csv").exists()
    assert (outdir / "cmd_update_excel.txt").exists()


def test_main_local_only_runs_local_artifacts(tmp_path, monkeypatch, capsys):
    tickets_path, arrivee_path, outdir, excel_path = _write_sample_payload(tmp_path)

    monkeypatch.setattr(drive_sync, "is_gcs_enabled", lambda: True)
    monkeypatch.setattr(
        drive_sync,
        "_build_service",
        MagicMock(side_effect=AssertionError("_build_service should be skipped")),
    )
    monkeypatch.setattr(
        drive_sync,
        "push_tree",
        MagicMock(side_effect=AssertionError("push_tree should not run")),
    )
    monkeypatch.setattr(
        drive_sync,
        "upload_file",
        MagicMock(side_effect=AssertionError("upload_file should not run")),
    )
    monkeypatch.setattr(
        drive_sync,
        "download_file",
        MagicMock(side_effect=AssertionError("download_file should not run")),
    )
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "drive_sync.py",
            "--local-only",
            "--arrivee",
            str(arrivee_path),
            "--tickets",
            str(tickets_path),
            "--outdir",
            str(outdir),
            "--excel",
            str(excel_path),
            "--places",
            "2",
        ],
    )

    result = drive_sync.main()
    captured = capsys.readouterr()

    assert result == 0
    assert "--local-only" in captured.out

    updated = json.loads(tickets_path.read_text(encoding="utf-8"))
    assert "roi_reel" in updated
    assert "brier_total" in updated

    arrivee_payload = json.loads((outdir / "arrivee.json").read_text(encoding="utf-8"))
    assert arrivee_payload["arrivee"]["result"] == ["1", "5"]

    csv_content = (outdir / "ligne_resultats.csv").read_text(encoding="utf-8")
    assert "ROI_reel" in csv_content.splitlines()[0]

    cmd = (outdir / "cmd_update_excel.txt").read_text(encoding="utf-8")
    assert "update_excel_with_results.py" in cmd

    assert excel_path.exists()
    workbook = load_workbook(excel_path)
    assert "ROI Observé" in workbook.sheetnames


def test_build_service_env(monkeypatch):
    creds_data = {"type": "service_account"}
    encoded = base64.b64encode(json.dumps(creds_data).encode()).decode()
    monkeypatch.setenv("GCS_SERVICE_KEY_B64", encoded)
    monkeypatch.setenv("GOOGLE_CLOUD_PROJECT", "proj-id")

    with patch(
        "scripts.drive_sync.service_account.Credentials.from_service_account_info"
    ) as cred_mock, patch("scripts.drive_sync.storage.Client") as client_mock:
        cred_mock.return_value = "creds"
        client_mock.return_value = "client"
        client = drive_sync._build_service()

    cred_mock.assert_called_once_with(creds_data, scopes=drive_sync.SCOPES)
    client_mock.assert_called_once_with(project="proj-id", credentials="creds")
    assert client == "client"


@pytest.mark.parametrize("project_env", [None, ""])
def test_build_service_missing_env(monkeypatch, project_env):
    monkeypatch.delenv("GCS_SERVICE_KEY_B64", raising=False)
    monkeypatch.delenv("GCS_SERVICE_KEY_JSON", raising=False)
    if project_env is None:
        monkeypatch.delenv("GOOGLE_CLOUD_PROJECT", raising=False)
    else:
        monkeypatch.setenv("GOOGLE_CLOUD_PROJECT", project_env)
    with patch("scripts.drive_sync.storage.Client") as client_mock:
        client = drive_sync._build_service()
    client_mock.assert_called_once_with(project=None)
    assert client == client_mock.return_value


def test_build_drive_service_from_file(monkeypatch, tmp_path):
    creds_path = tmp_path / "drive-creds.json"
    creds_path.write_text(json.dumps({"type": "service_account"}), encoding="utf-8")

    fake_creds = MagicMock(name="creds")
    builder = MagicMock(name="builder", return_value="service")

    monkeypatch.setattr(drive_sync, "_DRIVE_BUILD", builder)
    monkeypatch.setattr(drive_sync, "_MEDIA_FILE_UPLOAD", object())
    monkeypatch.setattr(drive_sync, "_MEDIA_DOWNLOAD", object())
    monkeypatch.setattr(
        drive_sync.service_account.Credentials,
        "from_service_account_file",
        MagicMock(return_value=fake_creds),
    )

    service = drive_sync._build_drive_service(creds_path)

    assert service == "service"
    drive_sync.service_account.Credentials.from_service_account_file.assert_called_once_with(
        str(creds_path), scopes=drive_sync.DRIVE_SCOPES
    )
    builder.assert_called_once_with(
        "drive", "v3", credentials=fake_creds, cache_discovery=False
    )


def test_build_drive_service_from_json(monkeypatch):
    payload = {"type": "service_account"}
    fake_creds = MagicMock(name="creds")
    builder = MagicMock(name="builder", return_value="service")

    monkeypatch.setattr(drive_sync, "_DRIVE_BUILD", builder)
    monkeypatch.setattr(drive_sync, "_MEDIA_FILE_UPLOAD", object())
    monkeypatch.setattr(drive_sync, "_MEDIA_DOWNLOAD", object())
    monkeypatch.setattr(
        drive_sync.service_account.Credentials,
        "from_service_account_info",
        MagicMock(return_value=fake_creds),
    )

    service = drive_sync._build_drive_service(json.dumps(payload))

    assert service == "service"
    drive_sync.service_account.Credentials.from_service_account_info.assert_called_once_with(
        payload, scopes=drive_sync.DRIVE_SCOPES
    )
    builder.assert_called_once_with(
        "drive", "v3", credentials=fake_creds, cache_discovery=False
    )


def test_drive_download_file(monkeypatch, tmp_path):
    content = b"drive-bytes"

    class FakeDownloader:
        def __init__(self, buffer, request):
            self.buffer = buffer
            self.request = request
            self._done = False

        def next_chunk(self):
            if self._done:
                return None, True
            self.buffer.write(content)
            self._done = True
            return None, True

    files_resource = MagicMock()
    files_resource.get_media.return_value = MagicMock(name="media-request")
    service = MagicMock()
    service.files.return_value = files_resource

    monkeypatch.setattr(drive_sync, "_MEDIA_DOWNLOAD", FakeDownloader)
    monkeypatch.setattr(drive_sync, "_MEDIA_FILE_UPLOAD", object())
    monkeypatch.setattr(drive_sync, "_DRIVE_BUILD", MagicMock())

    dest = tmp_path / "download.bin"
    result = drive_sync.drive_download_file(service, "file123", dest)

    assert result == dest
    assert dest.read_bytes() == content
    files_resource.get_media.assert_called_once_with(fileId="file123")


def test_drive_upload_file_create(monkeypatch, tmp_path):
    local = tmp_path / "payload.json"
    local.write_text("{}", encoding="utf-8")

    created: dict[str, Any] = {}

    class FakeUpload:
        def __init__(self, filename, mimetype=None, resumable=False):
            created["filename"] = filename
            created["mimetype"] = mimetype
            created["resumable"] = resumable

    files_resource = MagicMock()
    files_resource.create.return_value.execute.return_value = {"id": "new-id"}
    service = MagicMock()
    service.files.return_value = files_resource

    monkeypatch.setattr(drive_sync, "_MEDIA_FILE_UPLOAD", FakeUpload)
    monkeypatch.setattr(drive_sync, "_MEDIA_DOWNLOAD", object())
    monkeypatch.setattr(drive_sync, "_DRIVE_BUILD", MagicMock())

    result = drive_sync.drive_upload_file(service, "folder123", local)

    assert result == "new-id"
    files_resource.create.assert_called_once()
    kwargs = files_resource.create.call_args.kwargs
    assert kwargs["body"] == {"name": local.name, "parents": ["folder123"]}
    assert created["filename"] == str(local)
    assert created["resumable"] is False


def test_drive_upload_file_update(monkeypatch, tmp_path):
    local = tmp_path / "workbook.xlsx"
    local.write_bytes(b"data")

    class FakeUpload:
        def __init__(self, filename, mimetype=None, resumable=False):
            self.filename = filename
            self.mimetype = mimetype
            self.resumable = resumable

    files_resource = MagicMock()
    files_resource.update.return_value.execute.return_value = {"id": "file-id"}
    service = MagicMock()
    service.files.return_value = files_resource

    monkeypatch.setattr(drive_sync, "_MEDIA_FILE_UPLOAD", FakeUpload)
    monkeypatch.setattr(drive_sync, "_MEDIA_DOWNLOAD", object())
    monkeypatch.setattr(drive_sync, "_DRIVE_BUILD", MagicMock())

    result = drive_sync.drive_upload_file(
        service, None, local, file_id="file-id", mime_type="mime/test"
    )

    assert result == "file-id"
    files_resource.update.assert_called_once_with(fileId="file-id", media_body=ANY)
    upload_obj = files_resource.update.call_args.kwargs["media_body"]
    assert isinstance(upload_obj, FakeUpload)
    assert upload_obj.filename == str(local)
    assert upload_obj.mimetype == "mime/test"


def test_main_drive_flow(monkeypatch, tmp_path):
    tickets_path, arrivee_path, outdir, excel_path = _write_sample_payload(tmp_path)

    drive_service = MagicMock(name="drive_service")
    download_mock = MagicMock(name="download")
    upload_mock = MagicMock(name="upload")

    monkeypatch.setattr(
        drive_sync, "_build_drive_service", MagicMock(return_value=drive_service)
    )
    monkeypatch.setattr(drive_sync, "drive_download_file", download_mock)
    monkeypatch.setattr(drive_sync, "drive_upload_file", upload_mock)
    monkeypatch.setattr(drive_sync, "is_gcs_enabled", lambda: False)
    monkeypatch.setattr(drive_sync, "_build_service", MagicMock())
    monkeypatch.setattr(drive_sync, "push_tree", MagicMock())
    monkeypatch.setattr(drive_sync, "upload_file", MagicMock())
    monkeypatch.setattr(drive_sync, "download_file", MagicMock())

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "drive_sync.py",
            "--arrivee",
            str(arrivee_path),
            "--tickets",
            str(tickets_path),
            "--outdir",
            str(outdir),
            "--excel",
            str(excel_path),
            "--folder-id",
            "folder123",
            "--drive-credentials",
            str(tmp_path / "unused.json"),
            "--excel-file-id",
            "excel123",
            "--upload-result",
            "--upload-line",
        ],
    )

    result = drive_sync.main()

    assert result == 0
    download_mock.assert_called_once()
    excel_calls = [
        call
        for call in upload_mock.call_args_list
        if call.kwargs.get("file_id") == "excel123"
    ]
    assert len(excel_calls) == 1
    uploaded_names = {Path(call.args[2]).name for call in upload_mock.call_args_list}
    assert "arrivee.json" in uploaded_names
    assert "ligne_resultats.csv" in uploaded_names
