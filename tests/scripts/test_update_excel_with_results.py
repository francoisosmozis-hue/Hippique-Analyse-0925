# tests/scripts/test_update_excel_with_results.py
import pytest
from openpyxl import Workbook
from hippique_orchestrator.scripts.update_excel_with_results import (
    _as_float,
    _normalise_notes,
    _upsert_row,
    PREVISION_HEADERS,
)


def test_as_float():
    assert _as_float(10.5) == 10.5
    assert _as_float("25.5") == 25.5
    assert _as_float("invalid") == 0.0
    assert _as_float(None) == 0.0


def test_normalise_notes():
    assert _normalise_notes(["note1", "note2"]) == "note1; note2"
    assert _normalise_notes("single note") == "single note"
    assert _normalise_notes(None) == ""
    assert _normalise_notes(["note1", None, "note2"]) == "note1; note2"


def test_upsert_row():
    wb = Workbook()
    ws = wb.active

    # 1. Insert a new row
    headers = ["R/C", "hippodrome", "mises"]
    values1 = {"R/C": "R1C1", "hippodrome": "Vincennes", "mises": 10.0}

    row_idx = _upsert_row(ws, headers, values1)

    assert row_idx == 2
    assert ws.cell(row=1, column=1).value == "R/C"
    assert ws.cell(row=1, column=3).value == "mises"
    assert ws.cell(row=2, column=1).value == "R1C1"
    assert ws.cell(row=2, column=3).value == 10.0

    # 2. Update the existing row
    values2 = {"R/C": "R1C1", "hippodrome": "Vincennes", "mises": 15.5}
    row_idx_update = _upsert_row(ws, headers, values2)

    assert row_idx_update == 2  # Should be the same row
    assert ws.max_row == 2  # No new row should be added
    assert ws.cell(row=2, column=3).value == 15.5  # Value should be updated

    # 3. Insert a second row
    values3 = {"R/C": "R1C2", "hippodrome": "Auteuil", "mises": 5.0}
    row_idx_3 = _upsert_row(ws, headers, values3)
    assert row_idx_3 == 3
    assert ws.max_row == 3
    assert ws.cell(row=3, column=2).value == "Auteuil"
