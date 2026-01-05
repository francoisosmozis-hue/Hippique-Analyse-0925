
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from hippique_orchestrator.scripts.update_excel_with_results import update_excel


@pytest.fixture
def sample_payload(tmp_path):
    """Create a sample post-course payload JSON file."""
    payload = {
        "meta": {
            "rc": "R1C1",
            "hippodrome": "Vincennes",
            "date": "2023-01-01",
            "discipline": "Attelé",
            "model": "TestModel-v1"
        },
        "mises": {
            "total": 10.0,
            "gains": 15.0
        },
        "ev_estimees": {
            "roi_global": 0.25,
            "combined_expected_payout": 12.5
        },
        "ev_observees": {
            "verdict": "OK"
        },
        "notes": ["Test note 1", "Test note 2"],
        "tickets": []
    }
    payload_path = tmp_path / "payload.json"
    payload_path.write_text(json.dumps(payload))
    return str(payload_path)


def test_update_creates_new_excel_and_sheets(tmp_path, sample_payload):
    """
    Test that a new Excel file and its sheets are created correctly
    if the workbook does not exist.
    """
    excel_path = tmp_path / "new_workbook.xlsx"

    # Action
    update_excel(excel_path_str=str(excel_path), payload_path_str=sample_payload)

    # Verification
    assert excel_path.exists()

    # Load the created workbook and verify its contents
    wb = load_workbook(excel_path)

    # Check sheet names
    assert "Suivi" in wb.sheetnames
    assert "ROI Prévisionnel" in wb.sheetnames
    assert "ROI Observé" in wb.sheetnames

    # Check "Suivi" sheet
    ws_suivi = wb["Suivi"]
    assert ws_suivi.cell(row=1, column=1).value == "R/C"
    assert ws_suivi.max_row == 2
    assert ws_suivi.cell(row=2, column=1).value == "R1C1"
    assert ws_suivi.cell(row=2, column=5).value == 10.0  # Mises
    assert ws_suivi.cell(row=2, column=6).value == 15.0  # Gains
    assert ws_suivi.cell(row=2, column=7).value == 0.5   # ROI_reel
    assert ws_suivi.cell(row=2, column=11).value == "Test note 1; Test note 2" # Notes

    # Check "ROI Prévisionnel" sheet
    ws_prev = wb["ROI Prévisionnel"]
    assert ws_prev.cell(row=1, column=1).value == "R/C"
    assert ws_prev.max_row == 2
    assert ws_prev.cell(row=2, column=1).value == "R1C1"
    assert ws_prev.cell(row=2, column=9).value == 0.25 # ROI_global
    assert ws_prev.cell(row=2, column=14).value == "TestModel-v1" # model


def test_update_upserts_existing_row(tmp_path, sample_payload):
    """
    Test that an existing row in the Excel sheet is updated (upserted)
    instead of creating a new one.
    """
    excel_path = tmp_path / "existing_workbook.xlsx"

    # 1. Create a pre-existing workbook with one row
    wb = Workbook() # New workbook
    ws = wb.active
    ws.title = "Suivi"
    # Headers
    ws.cell(row=1, column=1, value="R/C")
    ws.cell(row=1, column=5, value="mises")
    # Old data
    ws.cell(row=2, column=1, value="R1C1")
    ws.cell(row=2, column=5, value=999) # Old stake value
    wb.save(excel_path)
    
    # Action: Run update_excel on the existing file
    update_excel(excel_path_str=str(excel_path), payload_path_str=sample_payload)
    
    # Verification
    wb_updated = load_workbook(excel_path)
    ws_updated = wb_updated["Suivi"]
    
    # Check that no new row was added
    assert ws_updated.max_row == 2 
    
    # Check that the existing row was updated
    assert ws_updated.cell(row=2, column=5).value == 10.0 # New stake value
