import json
from pathlib import Path

from openpyxl import load_workbook

from scripts import update_excel_planning as planner


def _read_planning_rows(excel_path: Path):
    wb = load_workbook(excel_path)
    ws = wb["Planning"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(values):
            continue
        rows.append(dict(zip(headers, values, strict=False)))
    return rows


def test_update_excel_planning_h30_then_h5(tmp_path):
    excel_path = tmp_path / "modele.xlsx"

    meeting_payload = {
        "date": "2024-09-25",
        "reunion": "R4",
        "course": "C5",
        "hippodrome": "Vincennes",
        "start_time": "13:45",
        "partants": 14,
        "discipline": "Trot",
    }
    meeting_json = tmp_path / "meeting.json"
    meeting_json.write_text(json.dumps(meeting_payload), encoding="utf-8")

    planner.main(
        [
            "--phase",
            "H30",
            "--in",
            str(meeting_json),
            "--excel",
            str(excel_path),
        ]
    )

    rows = _read_planning_rows(excel_path)
    assert len(rows) == 1
    row = rows[0]
    assert row["Date"] == "2024-09-25"
    assert row["Réunion"] == "R4"
    assert row["Course"] == "C5"
    assert row["Hippodrome"] == "Vincennes"
    assert row["Heure"] == "13:45"
    assert row["Partants"] == 14
    assert row["Discipline"] == "Trot"
    assert row["Statut H-30"] == "Collecté"
    assert row.get("Statut H-5") in (None, "")

    course_dir = tmp_path / "R4C5"
    course_dir.mkdir()
    h5_payload = {
        "date": "2024-09-25",
        "reunion": "R4",
        "course": "C5",
        "hippodrome": "Vincennes",
        "start_time": "13:45",
        "partants": 14,
        "discipline": "Trot",
        "tickets": [
            {
                "type": "SP",
                "legs": [{"selections": [3, 5]}],
                "odds": 2.0,
            },
            {
                "type": "CPL",
                "legs": [{"horses": [1, 3]}],
                "odds": 1.5,
            },
        ],
    }
    (course_dir / "analysis_H5.json").write_text(
        json.dumps(h5_payload), encoding="utf-8"
    )

    planner.main(
        [
            "--phase",
            "H5",
            "--in",
            str(course_dir),
            "--excel",
            str(excel_path),
        ]
    )

    rows = _read_planning_rows(excel_path)
    assert len(rows) == 1
    row = rows[0]
    assert row["Statut H-30"] == "Collecté"
    assert row["Statut H-5"] == "Analysé"
    assert row["Jouable H-5"] == "Oui"
    assert row["Tickets H-5"] == "SP:3-5@2 | CPL:1-3@1.5"
    assert row["Heure"] == "13:45"
    assert row["Partants"] == 14
