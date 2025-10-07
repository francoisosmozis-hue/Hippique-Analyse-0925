import json

from openpyxl import load_workbook

from scripts import update_excel_planning as updater


def _headers(ws):
    return {
        str(ws.cell(row=1, column=col).value): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=col).value not in (None, "")
    }


def test_update_excel_h30_inserts_rows(tmp_path):
    workbook = tmp_path / "planning.xlsx"
    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()

    snapshot = {
        "date": "2023-09-25",
        "reunion": "R1",
        "course": "C3",
        "hippodrome": "Paris-Vincennes",
        "start_time": "13:45",
        "partants": 12,
        "discipline": "TROT",
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(snapshot), encoding="utf-8")

    updater.main(
        [
            "--phase",
            "H30",
            "--in",
            str(meeting_dir),
            "--excel",
            str(workbook),
        ]
    )

    wb = load_workbook(workbook)
    ws = wb["Planning"]
    headers = _headers(ws)

    assert ws.cell(row=2, column=headers["Date"]).value == "2023-09-25"
    assert ws.cell(row=2, column=headers["Réunion"]).value == "R1"
    assert ws.cell(row=2, column=headers["Course"]).value == "C3"
    assert ws.cell(row=2, column=headers["Hippodrome"]).value == "Paris-Vincennes"
    assert ws.cell(row=2, column=headers["Heure"]).value == "13:45"
    assert ws.cell(row=2, column=headers["Partants"]).value == 12
    assert ws.cell(row=2, column=headers["Discipline"]).value == "TROT"
    assert ws.cell(row=2, column=headers["Statut H-30"]).value == "Collecté"


def test_update_excel_h5_updates_existing_row(tmp_path):
    workbook = tmp_path / "planning.xlsx"

    meeting_dir = tmp_path / "meeting"
    meeting_dir.mkdir()
    snapshot = {
        "date": "2023-09-25",
        "reunion": "R1",
        "course": "C3",
        "hippodrome": "Paris-Vincennes",
        "start_time": "13:45",
        "partants": 12,
        "discipline": "TROT",
    }
    (meeting_dir / "snapshot.json").write_text(json.dumps(snapshot), encoding="utf-8")

    updater.main(
        [
            "--phase",
            "H30",
            "--in",
            str(meeting_dir),
            "--excel",
            str(workbook),
        ]
    )

    analysis_dir = tmp_path / "R1C3"
    analysis_dir.mkdir()
    analysis_payload = {
        "date": "2023-09-25",
        "reunion": "R1",
        "course": "C3",
        "hippodrome": "Paris-Vincennes",
        "start_time": "13:45",
        "partants": 12,
        "discipline": "TROT",
        "abstain": False,
        "roi": 0.35,
        "tickets": [
            {
                "type": "SP",
                "legs": [{"selections": [3, 5]}],
                "odds": 2.0,
            },
            {
                "type": "CPL",
                "legs": [{"selections": [1, 3]}],
                "odds": 1.8,
            },
        ],
    }
    (analysis_dir / "analysis_H5.json").write_text(
        json.dumps(analysis_payload), encoding="utf-8"
    )

    updater.main(
        [
            "--phase",
            "H5",
            "--in",
            str(analysis_dir),
            "--excel",
            str(workbook),
        ]
    )

    wb = load_workbook(workbook)
    ws = wb["Planning"]
    headers = _headers(ws)

    assert ws.cell(row=2, column=headers["Statut H-5"]).value == "Analysé"
    assert ws.cell(row=2, column=headers["Jouable H-5"]).value == "Oui"
    assert (
        ws.cell(row=2, column=headers["Tickets H-5"]).value == "SP:3-5@2 | CPL:1-3@1.8"
    )
    assert ws.cell(row=2, column=headers["Commentaires"]).value == "ROI estimé 35%"
