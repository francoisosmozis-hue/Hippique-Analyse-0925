"""Utilities for synchronising the planning worksheet with pipeline outputs."""

from __future__ import annotations

import argparse
import datetime as dt
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


PLANNING_HEADERS: Sequence[str] = (
    "Date",
    "Réunion",
    "Course",
    "Hippodrome",
    "Heure",
    "Partants",
    "Discipline",
    "Statut H-30",
    "Statut H-5",
    "Jouable H-5",
    "Tickets H-5",
    "Commentaires",
)


def _phase_argument(value: str) -> str:
    cleaned = value.strip().upper().replace("-", "")
    if cleaned not in {"H30", "H5"}:
        raise argparse.ArgumentTypeError("phase must be H30 or H5")
    return cleaned


def _load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def _load_workbook(path: Path) -> tuple[Workbook, bool]:
    if path.exists():
        return load_workbook(path), False
    return Workbook(), True


def _prepare_sheet(wb: Workbook, title: str) -> Worksheet:
    if title in wb.sheetnames:
        return wb[title]
    if len(wb.sheetnames) == 1:
        ws = wb.active
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value in (None, ""):
            ws.title = title
            return ws
    return wb.create_sheet(title=title)


def _ensure_headers(ws: Worksheet, headers: Sequence[str]) -> Mapping[str, int]:
    header_map: Dict[str, int] = {}
    blank_sheet = (
        ws.max_row <= 1
        and ws.max_column == 1
        and ws.cell(row=1, column=1).value in (None, "")
    )
    max_col = 0 if blank_sheet else ws.max_column
    if not blank_sheet:
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value not in (None, ""):
                header_map[str(value)] = col
    next_col = max_col
    for header in headers:
        if header not in header_map:
            next_col += 1
            ws.cell(row=1, column=next_col, value=header)
            header_map[header] = next_col
    return header_map


def _coerce_number(value: Any) -> str:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return str(value)
    if abs(num - round(num)) < 1e-9:
        return str(int(round(num)))
    return f"{num:.2f}".rstrip("0").rstrip(".")


def _format_time(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    if len(text) == 4 and text.isdigit():
        return f"{text[:2]}:{text[2:]}"
    if len(text) == 5 and text[2] == ":":
        return text
    cleaned = text.replace("Z", "+00:00")
    try:
        parsed = dt.datetime.fromisoformat(cleaned)
    except ValueError:
        return text
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(dt.timezone.utc)
    return parsed.strftime("%H:%M")


def _first_non_empty(*values: Any) -> Any:
    for value in values:
        if isinstance(value, str):
            trimmed = value.strip()
            if trimmed:
                return trimmed
        elif value not in (None, ""):
            return value
    return None


def _blank_if_missing(value: Any) -> Any:
    return "" if value in (None, "") else value


def _extract_common_meta(payload: Mapping[str, Any]) -> Dict[str, Any]:
    meta = payload.get("meta") if isinstance(payload.get("meta"), Mapping) else {}
    date = _first_non_empty(meta.get("date"), payload.get("date"))
    hippodrome = _first_non_empty(meta.get("hippodrome"), meta.get("hippo"), payload.get("hippodrome"), payload.get("hippo"))
    reunion = _first_non_empty(
        meta.get("reunion"),
        meta.get("meeting"),
        meta.get("r"),
        payload.get("reunion"),
        payload.get("meeting"),
        payload.get("r"),
        payload.get("r_label"),
    )
    course = _first_non_empty(
        meta.get("course"),
        meta.get("race"),
        meta.get("c"),
        payload.get("course"),
        payload.get("race"),
        payload.get("c"),
        payload.get("course_label"),
    )
    partants = _first_non_empty(
        meta.get("partants"),
        payload.get("partants"),
        meta.get("nb_partants"),
        payload.get("nb_partants"),
        meta.get("runners_count"),
        payload.get("runners_count"),
    )
    if partants in (None, ""):
        runners = meta.get("runners") if isinstance(meta, Mapping) else None
        if not runners:
            runners = payload.get("runners")
        if isinstance(runners, Iterable) and not isinstance(runners, (str, bytes, Mapping)):
            partants = sum(1 for _ in runners)
    start_time = _first_non_empty(
        meta.get("start_time"),
        meta.get("start"),
        meta.get("heure"),
        meta.get("time"),
        payload.get("start_time"),
        payload.get("start"),
        payload.get("heure"),
        payload.get("time"),
        payload.get("hour"),
    )
    formatted_time = _format_time(start_time)
    discipline = _first_non_empty(
        meta.get("discipline"),
        meta.get("type"),
        payload.get("discipline"),
        payload.get("type"),
    )
    return {
        "date": date,
        "hippodrome": hippodrome,
        "reunion": reunion,
        "course": course,
        "start_time": formatted_time,
        "partants": partants,
        "discipline": discipline,
    }


def _upsert(ws: Worksheet, header_map: Mapping[str, int], row: Mapping[str, Any], keys: Sequence[str]) -> None:
    key_values = {key: row.get(key) for key in keys}
    for key, value in key_values.items():
        if value in (None, ""):
            raise ValueError(f"Missing key value for {key}")

    target_row = None
    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        if all(str(ws.cell(row=row_idx, column=header_map[key]).value or "") == str(key_values[key]) for key in keys):
            target_row = row_idx
            break
    if target_row is None:
        target_row = max(2, max_row + 1)
    for header, value in row.items():
        if header not in header_map or value is None:
            continue
        ws.cell(row=target_row, column=header_map[header], value=value)


def _summarise_tickets(tickets: Any) -> str:
    if not isinstance(tickets, Iterable):
        return ""
    summaries: List[str] = []
    for ticket in tickets:
        if not isinstance(ticket, Mapping):
            continue
        label = _first_non_empty(ticket.get("type"), ticket.get("bet_type"), ticket.get("id"))
        legs = ticket.get("legs")
        horses: List[str] = []
        if isinstance(legs, Iterable) and not isinstance(legs, (str, bytes)):
            for leg in legs:
                if isinstance(leg, Mapping):
                    selections = leg.get("selections")
                    if isinstance(selections, Iterable) and not isinstance(selections, (str, bytes)):
                        selection_text = "-".join(str(x) for x in selections if x not in (None, ""))
                    else:
                        selection_text = None
                    horses_list = leg.get("horses")
                    if isinstance(horses_list, Iterable) and not isinstance(horses_list, (str, bytes)):
                        horses_text = "-".join(str(x) for x in horses_list if x not in (None, ""))
                    else:
                        horses_text = None
                    horse = _first_non_empty(
                        leg.get("horse"),
                        leg.get("selection"),
                        leg.get("id"),
                        leg.get("num"),
                        leg.get("number"),
                        selection_text,
                        horses_text,
                    )
                else:
                    horse = leg
                if horse not in (None, ""):
                    horses.append(str(horse))
        odds = _first_non_empty(ticket.get("odds"), ticket.get("rapport"), ticket.get("payout"))
        horses_joined = "-".join(horses)
        segment = ""
        if label:
            segment = str(label)
        if horses_joined:
            segment = f"{segment}:{horses_joined}" if segment else horses_joined
        if odds not in (None, ""):
            segment = f"{segment}@{_coerce_number(odds)}"
        stake = _first_non_empty(ticket.get("stake"), ticket.get("mise"), ticket.get("amount"))
        if stake not in (None, ""):
            segment = f"{segment}({_coerce_number(stake)}€)"
        if segment:
            summaries.append(segment)
    return " | ".join(summaries)


def _extract_roi(payload: Mapping[str, Any]) -> float | None:
    candidates = [
        payload.get("roi"),
        payload.get("roi_est"),
    ]
    validation = payload.get("validation")
    if isinstance(validation, Mapping):
        candidates.extend(
            [
                validation.get("roi_global_est"),
                (validation.get("sp") or {}).get("roi_est") if isinstance(validation.get("sp"), Mapping) else None,
            ]
        )
    for candidate in candidates:
        try:
            return float(candidate)
        except (TypeError, ValueError):
            continue
    return None


def _status_h30(default: str) -> str:
    return default


def _status_h5() -> str:
    return "Analysé"


def _jouable_flag(payload: Mapping[str, Any]) -> str:
    return "Non" if bool(payload.get("abstain")) else "Oui"


def _comment_h5(payload: Mapping[str, Any]) -> str:
    abstain = bool(payload.get("abstain"))
    reason = _first_non_empty(
        payload.get("abstain_reason"),
        payload.get("notes"),
        payload.get("message"),
    )
    roi = _extract_roi(payload)
    if abstain:
        return reason or ""
    if roi is not None:
        return f"ROI {roi:.2f}"
    return reason or ""


def _collect_h30_entries(source: Path, status: str) -> List[Dict[str, Any]]:
    entries: List[Dict[str, Any]] = []
    if source.is_file():
        payloads = [(source, _load_json(source))]
    elif source.is_dir():
        payloads = [(path, _load_json(path)) for path in sorted(source.rglob("*.json"))]
    else:
        raise FileNotFoundError(source)
    for path, payload in payloads:
        if not isinstance(payload, Mapping):
            continue
        meta = _extract_common_meta(payload)
        if not (meta.get("date") and meta.get("reunion") and meta.get("course")):
            continue
        row = {
            "Date": meta.get("date"),
            "Réunion": meta.get("reunion"),
            "Course": meta.get("course"),
            "Hippodrome": _blank_if_missing(meta.get("hippodrome")),
            "Heure": _blank_if_missing(meta.get("start_time")),
            "Partants": _blank_if_missing(meta.get("partants")),
            "Discipline": _blank_if_missing(meta.get("discipline")),
            "Statut H-30": _status_h30(status),
            "Commentaires": "",
        }
        entries.append(row)
    return entries


def _load_h5_payload(source: Path) -> Mapping[str, Any]:
    if source.is_file():
        return _load_json(source)
    if source.is_dir():
        for name in ("analysis_H5.json", "analysis.json"):
            candidate = source / name
            if candidate.is_file():
                return _load_json(candidate)
    raise FileNotFoundError(source)


def _prepare_h5_row(payload: Mapping[str, Any]) -> Dict[str, Any]:
    meta = _extract_common_meta(payload)
    tickets = payload.get("tickets") if isinstance(payload, Mapping) else None
    summary = _summarise_tickets(tickets)
    row: Dict[str, Any] = {
        "Date": meta.get("date"),
        "Réunion": meta.get("reunion"),
        "Course": meta.get("course"),
        "Hippodrome": _blank_if_missing(meta.get("hippodrome")),
        "Heure": _blank_if_missing(meta.get("start_time")),
        "Partants": _blank_if_missing(meta.get("partants")),
        "Discipline": _blank_if_missing(meta.get("discipline")),
        "Statut H-5": _status_h5(),
        "Jouable H-5": _jouable_flag(payload)
    }
    row["Tickets H-5"] = summary or ""
    row["Commentaires"] = _comment_h5(payload) or ""
    return row


def main(argv: Sequence[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Met à jour l'onglet Planning de l'Excel")
    parser.add_argument("--phase", required=True, type=_phase_argument, help="Phase à appliquer (H30/H5)")
    parser.add_argument("--input", "--in", dest="source", required=True, help="Fichier ou dossier source")
    parser.add_argument("--excel", required=True, help="Chemin du classeur Excel")
    parser.add_argument("--sheet", default="Planning", help="Nom de l'onglet à mettre à jour")
    parser.add_argument(
        "--status-h30",
        default="Collecté",
        help="Statut par défaut renseigné pour les lignes H-30",
    )
    parser.add_argument("--drive-folder", help=argparse.SUPPRESS)
    args = parser.parse_args(argv)

    source = Path(args.source)
    excel_path = Path(args.excel)
    wb, _ = _load_workbook(excel_path)
    ws = _prepare_sheet(wb, args.sheet)
    header_map = _ensure_headers(ws, PLANNING_HEADERS)

    if args.phase == "H30":
        rows = _collect_h30_entries(source, status=args.status_h30)
        keys = ("Date", "Réunion", "Course")
        for row in rows:
            _upsert(ws, header_map, row, keys)
        message = f"{len(rows)} ligne(s) H-30 mises à jour"
    else:
        payload = _load_h5_payload(source)
        row = _prepare_h5_row(payload)
        keys = ("Date", "Réunion", "Course")
        if not all(row.get(key) for key in keys):
            raise ValueError("Impossible de déterminer Date/Réunion/Course pour la mise à jour H-5")
        _upsert(ws, header_map, row, keys)
        message = "1 ligne H-5 mise à jour"

    wb.save(excel_path)
    print(message)


if __name__ == "__main__":  # pragma: no cover
    main()
