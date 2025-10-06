"""Update the tracking Excel workbook with projected and observed ROI."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from post_course_payload import (
    JsonDict,
    PostCourseSummary,
    build_payload_from_sources,
    merge_meta,
    summarise_ticket_metrics,
)


PREVISION_HEADERS = [
    "R/C",
    "hippodrome",
    "date",
    "discipline",
    "mises",
    "EV_sp",
    "EV_global",
    "ROI_sp",
    "ROI_global",
    "risk_of_ruin",
    "clv_moyen",
    "variance",
    "payout_attendu",
    "model",
]


OBSERVED_HEADERS = [
    "R/C",
    "hippodrome",
    "date",
    "discipline",
    "mises",
    "gains",
    "ROI_reel",
    "result_moyen",
    "ROI_reel_moyen",
    "Brier_total",
    "Brier_moyen",
    "EV_total",
    "EV_ecart",
    "model",
]


def _load_json(path: str | Path) -> JsonDict:
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def _load_workbook(path: Path) -> tuple[Workbook, bool]:
    if path.exists():
        return load_workbook(path), False
    return Workbook(), True


def _get_sheet(wb: Workbook, title: str) -> Worksheet:
    if title in wb.sheetnames:
        return wb[title]
    return wb.create_sheet(title=title)


def _ensure_header_map(ws: Worksheet, headers: Iterable[str]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    max_column = ws.max_column
    blank_sheet = (
        ws.max_row <= 1
        and ws.max_column == 1
        and ws.cell(row=1, column=1).value in (None, "")
    )
    if not blank_sheet:
        for col in range(1, max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value not in (None, ""):
                header_map[str(value)] = col
    next_col = 0 if blank_sheet else max_column
    for header in headers:
        if header not in header_map:
            next_col += 1
            ws.cell(row=1, column=next_col, value=header)
            header_map[header] = next_col
    return header_map


def _upsert_row(
    ws: Worksheet,
    headers: Iterable[str],
    values: dict[str, Any],
    *,
    key_header: str = "R/C",
) -> None:
    header_map = _ensure_header_map(ws, headers)
    key_value = values.get(key_header)
    if key_value in (None, ""):
        return
    key_col = header_map[key_header]
    key_str = str(key_value)
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=key_col).value
        if cell_value is None:
            continue
        if str(cell_value) == key_str:
            target_row = row_idx
            break
    if target_row is None:
        target_row = max(ws.max_row + 1, 2)
    for header in headers:
        col_idx = header_map[header]
        ws.cell(row=target_row, column=col_idx, value=values.get(header))


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _update_previsionnel_sheet(
    wb: Workbook,
    sheet_name: str,
    meta: JsonDict,
    *,
    total_stake: float,
    ev_data: JsonDict | None,
) -> None:
    rc = meta.get("rc")
    if not rc:
        return
    ws = _get_sheet(wb, sheet_name)
    ev_section = ev_data or {}
    row = {
        "R/C": rc,
        "hippodrome": meta.get("hippodrome", ""),
        "date": meta.get("date", ""),
        "discipline": meta.get("discipline", ""),
        "mises": round(total_stake, 2) if total_stake else 0.0,
        "EV_sp": ev_section.get("sp"),
        "EV_global": ev_section.get("global"),
        "ROI_sp": ev_section.get("roi_sp"),
        "ROI_global": ev_section.get("roi_global"),
        "risk_of_ruin": ev_section.get("risk_of_ruin"),
        "clv_moyen": ev_section.get("clv_moyen"),
        "variance": ev_section.get("variance"),
        "payout_attendu": ev_section.get("combined_expected_payout"),
        "model": meta.get("model") or meta.get("MODEL", ""),
    }
    _upsert_row(ws, PREVISION_HEADERS, row)


def _update_observe_sheet(
    wb: Workbook,
    sheet_name: str,
    meta: JsonDict,
    *,
    tickets: list[JsonDict],
    summary: PostCourseSummary,
    observed: JsonDict | None,
    total_stake: float,
    total_gain: float,
) -> None:
    rc = meta.get("rc")
    if not rc:
        return
    observed_section = observed or {}
    roi_observed = observed_section.get("roi_reel", summary.roi)
    if roi_observed is None and total_stake:
        roi_observed = (total_gain - total_stake) / total_stake if total_stake else 0.0
    has_observed_metrics = any(
        (
            roi_observed,
            total_gain,
            any("gain_reel" in t for t in tickets),
        )
    )
    if not has_observed_metrics:
        return

    row = {
        "R/C": rc,
        "hippodrome": meta.get("hippodrome", ""),
        "date": meta.get("date", ""),
        "discipline": meta.get("discipline", ""),
        "mises": round(total_stake, 2) if total_stake else 0.0,
        "gains": round(total_gain, 2) if total_gain else 0.0,
        "ROI_reel": roi_observed,
        "result_moyen": observed_section.get("result_moyen", summary.result_mean),
        "ROI_reel_moyen": observed_section.get(
            "roi_reel_moyen", summary.roi_ticket_mean
        ),
        "Brier_total": observed_section.get(
            "brier_total", summary.brier_total
        ),
        "Brier_moyen": observed_section.get("brier_moyen", summary.brier_mean),
        "EV_total": observed_section.get("ev_total", summary.ev_total),
        "EV_ecart": observed_section.get(
            "ev_ecart_total", summary.ev_diff_total
        ),
        "model": meta.get("model") or meta.get("MODEL", ""),
    }
    ws = _get_sheet(wb, sheet_name)
    _upsert_row(ws, OBSERVED_HEADERS, row)


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Mettre à jour le ROI dans l'Excel de suivi")
    parser.add_argument(
        "--excel",
        default="modele_suivi_courses_hippiques.xlsx",
        help="Chemin du classeur Excel à mettre à jour",
    )
    parser.add_argument("--payload", help="JSON normalisé généré après le post-course")
    parser.add_argument("--arrivee", help="JSON de l'arrivée officielle")
    parser.add_argument("--tickets", help="JSON des tickets (p_finale.json)")
    parser.add_argument(
        "--sheet-prevision",
        default="ROI Prévisionnel",
        help="Nom de la feuille pour les projections",
    )
    parser.add_argument(
        "--sheet-observe",
        default="ROI Observé",
        help="Nom de la feuille pour le ROI observé",
    )
    args = parser.parse_args(argv)

    excel_path = Path(args.excel)

    if not args.payload and (not args.arrivee or not args.tickets):
        parser.error("--payload ou le duo --arrivee/--tickets doit être fourni")

    if args.payload:
        payload_path = Path(args.payload)
        payload = _load_json(payload_path)
    else:
        arrivee_path = Path(args.arrivee)
        tickets_path = Path(args.tickets)
        arrivee_data = _load_json(arrivee_path)
        tickets_data = _load_json(tickets_path)
        payload = build_payload_from_sources(arrivee_data, tickets_data)

    if not isinstance(payload, dict):
        raise ValueError("Le payload post-course doit être un objet JSON.")

    meta_source = payload.get("meta", {}) if isinstance(payload, dict) else {}
    payload_arrivee = payload.get("arrivee") if isinstance(payload, dict) else {}
    meta = merge_meta(payload_arrivee or {}, {"meta": meta_source})

    raw_tickets = payload.get("tickets")
    tickets_list = [t for t in raw_tickets if isinstance(t, dict)] if isinstance(raw_tickets, list) else []
    summary = summarise_ticket_metrics(tickets_list)

    mises_raw = payload.get("mises")
    mises_section = mises_raw if isinstance(mises_raw, dict) else {}
    total_stake_value = mises_section.get("total") if mises_section else None
    if total_stake_value is None and mises_section:
        total_stake_value = mises_section.get("totales")
    total_stake = _as_float(total_stake_value, summary.total_stake)
    total_gain = _as_float(mises_section.get("gains") if mises_section else None, summary.total_gain)

    ev_raw = payload.get("ev_estimees")
    ev_data = ev_raw if isinstance(ev_raw, dict) else {}
    observed_raw = payload.get("ev_observees")
    observed = observed_raw if isinstance(observed_raw, dict) else {}

    wb, created = _load_workbook(excel_path)
    if created and "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        default_ws = wb["Sheet"]
        if default_ws.max_row <= 1 and default_ws.cell(row=1, column=1).value in (None, ""):
            wb.remove(default_ws)

    _update_previsionnel_sheet(
        wb,
        args.sheet_prevision,
        meta,
        total_stake=total_stake,
        ev_data=ev_data,
    )
    _update_observe_sheet(
        wb,
        args.sheet_observe,
        meta,
        tickets=tickets_list,
        summary=summary,
        observed=observed,
        total_stake=total_stake,
        total_gain=total_gain,
    )

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)


if __name__ == "__main__":  # pragma: no cover
    main()
